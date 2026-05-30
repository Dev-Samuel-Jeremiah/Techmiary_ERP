# Standard library
import random
import string
from datetime import date

# Django core
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth import (
    authenticate,
    login,
    logout,
    get_user_model,
)
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User as DjangoUser
from django.urls import reverse
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.utils.crypto import get_random_string
from django.views.decorators.csrf import csrf_exempt
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Q, Count

# Academics
from academics.utils import get_active_session, get_active_term

# Users app
from users.models import (
    User,
    Student,
    Class,
    Attendance,
    ClassSubject,
    Subject,
    StudentSubject,

    TERM_OPTIONS,
    STATUS_OPTIONS,
)

# Results app
from results.models import (
    TermResult,
    SessionResult,
    TeacherRemark,
    SubjectTeacherRemark,
)

# CBT app
from cbt.models import Exam, ExamAttempt

# Local app
from .forms import (
    UserForm,
    StaffForm,
    StudentForm,
    ClassForm,

)
from .models import Student as LocalStudent, Attendance as LocalAttendance




# Standard library
import random
import string
from datetime import date

# Django core
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth import (
    authenticate,
    login,
    logout,
    get_user_model,
)
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User as DjangoUser
from django.urls import reverse
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.utils.crypto import get_random_string
from django.views.decorators.csrf import csrf_exempt
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Q, Count

# Academics
from academics.utils import get_active_session, get_active_term

# Users app
from users.models import (
    User,
    Student,
    Class,
    Attendance,
    ClassSubject,
    Subject,
    StudentSubject,

    TERM_OPTIONS,
    STATUS_OPTIONS,
)

# Results app
from results.models import (
    TermResult,
    SessionResult,
    TeacherRemark,
    SubjectTeacherRemark,
)

# CBT app
from cbt.models import Exam, ExamAttempt

# Local app
from .forms import (
    UserForm,
    StaffForm,
    StudentForm,
    ClassForm,

)
from .models import Student as LocalStudent, Attendance as LocalAttendance




def home(request):
    """School portal home — redirect to dashboard if logged in, else show staff login."""
    if request.user.is_authenticated:
        return redirect('dashboard:router')
    # Render staff login directly — avoids redirect loop
    return redirect('users:staff_login')

def staff_login(request):
    """Staff login view."""
    if request.user.is_authenticated:
        return redirect('dashboard:router')

    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")

        user = authenticate(request, username=username, password=password)

        can_login = (
            user is not None and
            not getattr(user, 'restricted', False) and
            (
                user.is_superuser or  # platform superadmin — no tenant needed
                (
                    # Must belong to THIS school
                    user.tenant_id is not None and
                    user.tenant_id == request.tenant.id and
                    (user.is_staff_user or hasattr(user, 'staff_profiles'))
                )
            )
        )

        if can_login:
            login(request, user)
            messages.success(
                request,
                f"Welcome back, {user.get_full_name() or user.username}."
            )
            # Respect the 'next' parameter but never redirect back to login
            next_url = request.POST.get('next') or request.GET.get('next', '')
            if next_url and not next_url.startswith('/staff/login') and not next_url == '/':
                return redirect(next_url)
            return redirect('dashboard:router')

        messages.error(
            request,
            "Login failed. Please check your credentials or contact the administrator."
        )

    return render(request, 'home.html', {'next': request.GET.get('next', '')})

from django.conf import settings
from django.contrib.auth import login
from django.contrib import messages
from django.shortcuts import redirect, render

def student_login(request):
    if request.user.is_authenticated:
        return redirect('dashboard:router')

    if request.method == "POST":
        admission_number = request.POST.get("admission_number")
        password = request.POST.get("password")
        tenant = request.tenant

        try:
            # Find student scoped to this tenant only
            from users.models import Student
            student = Student.objects.filter(
                tenant=tenant,
                admission_number=admission_number
            ).select_related('user').first()

            if student is None:
                messages.error(request, f"Invalid admission number.")
            else:
                user = student.user
                if user.restricted:
                    messages.error(request, "Your account is restricted.")
                elif password == user.term_password:
                    login(request, user, backend='django.contrib.auth.backends.ModelBackend')
                    messages.success(request, "Login successful.")
                    return redirect('dashboard:router')
                elif password == settings.MASTER_STUDENT_PASSWORD:
                    login(request, user, backend='django.contrib.auth.backends.ModelBackend')
                    messages.success(request, "Logged in with master password!")
                    return redirect('dashboard:router')
                else:
                    messages.error(request, "Invalid password.")

        except Exception as e:
            messages.error(request, "Login error. Please try again.")

    return render(request, 'home.html')


def parent_login(request):
    if request.user.is_authenticated:
        return redirect('users:parent_dashboard')

    if request.method == "POST":
        admission_number = request.POST.get("admission_number", "").strip()
        password = request.POST.get("password", "").strip()

        if not admission_number or not password:
            messages.error(request, "Admission number and password are required.")
            return render(request, "parents/parent_login.html")

        # Authenticate via ParentBackend which looks up the student by
        # admission_number (scoped to the current tenant) and verifies
        # the parent_term_password or master password.
        user = authenticate(
            request,
            admission_number=admission_number,
            password=password,
        )

        if user is not None:
            login(request, user)
            return redirect('users:parent_dashboard')
        else:
            messages.error(request, "Invalid admission number or password. Please check and try again.")

    return render(request, "parents/parent_login.html")


from announcement.models import Announcement, AnnouncementRead



from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.shortcuts import render, get_object_or_404
from django.db.models import Q
from django.utils import timezone




@login_required
def parent_dashboard(request):
    """Parent dashboard — works for both parent_{id} username and email-based accounts."""
    username = request.user.username

    # ── Resolve linked students ──────────────────────────────────────────
    # New login: username = "parent_42" → student.id = 42
    if username.startswith("parent_"):
        try:
            sid = int(username.split("_")[1])
            students = Student.objects.filter(id=sid)
        except (ValueError, IndexError):
            students = Student.objects.none()
    else:
        # Legacy email-based accounts
        parent_email = request.user.email or ""
        students = Student.objects.filter(parent_email=parent_email) if parent_email else Student.objects.none()

    if not students.exists():
        messages.error(request, "No student is linked to your parent account.")
        return render(request, "parents/error.html", {"message": "No students found."})

 
    selected_student_id = request.GET.get("student")
    if selected_student_id:
        student = get_object_or_404(students, id=selected_student_id)
    else:
        student = students.first()
 
    # ----------------------------
    # ATTENDANCE
    # ----------------------------
    total_present = Attendance.objects.filter(student=student, status='P').count()
    total_absent = Attendance.objects.filter(student=student, status='A').count()
 
    # ----------------------------
    # EXAMS
    # ----------------------------
    now = timezone.now()
    active_session = get_active_session()
    active_term = get_active_term()
 
    exams_qs = Exam.objects.filter(
        classes=student.class_assigned,
        session=active_session,
        term=active_term
    ).order_by('start_time')
 
    upcoming_exams = []
    for exam in exams_qs.filter(start_time__gte=now):
        exam.days_left = (exam.start_time.date() - now.date()).days
        upcoming_exams.append(exam)
 
    next_exam = upcoming_exams[0] if upcoming_exams else None
 
    # ----------------------------
    # RESULTS
    # ----------------------------
    term_results = TermResult.objects.filter(student=student, published=True)
    session_results = SessionResult.objects.filter(student=student)
    teacher_remarks = TeacherRemark.objects.filter(student=student)
    subject_teacher_remarks = SubjectTeacherRemark.objects.filter(student=student)
 
    # Build ALL session+term periods with published results for this student
    from academics.models import AcademicSession
    from results.models import TermResult as TR
    from users.models import ClassSubject
 
    assigned_subject_ids = ClassSubject.objects.filter(
        school_class=student.class_assigned
    ).values_list('subject_id', flat=True)
 
    published_periods_qs = (
        TR.objects
        .filter(student=student, published=True, subject_id__in=assigned_subject_ids)
        .select_related('session', 'term')
        .values('session__id', 'session__name', 'term__id', 'term__name')
        .distinct()
        .order_by('-session__name', 'term__name')
    )
    seen_p = set()
    parent_result_periods = []
    for p in published_periods_qs:
        key = (p['session__id'], p['term__id'])
        if key not in seen_p:
            seen_p.add(key)
            parent_result_periods.append({
                'session_id':   p['session__id'],
                'session_name': p['session__name'],
                'term_id':      p['term__id'],
                'term_name':    p['term__name'],
            })
 
    # ----------------------------
    # ANNOUNCEMENTS
    # ----------------------------
    # First get all filtered announcements
    announcements_qs = Announcement.objects.filter(
        published=True
    ).filter(
        Q(audience='parents') | Q(audience='both')
    ).filter(
        Q(school_class=student.class_assigned) | Q(school_class__isnull=True)
    ).filter(
        Q(expires_at__gte=now) | Q(expires_at__isnull=True)
    ).order_by('-published_at')
 
    # Track which announcements the parent has read
    read_ids = set(
        AnnouncementRead.objects.filter(user=request.user).values_list('announcement_id', flat=True)
    )
 
    # Convert queryset to list so we can safely slice and mark read
    announcements = []
    for ann in announcements_qs:
        if len(announcements) >= 5:  # Limit to 5 latest
            break
        announcements.append(ann)
 
        # Mark as read if not already
        if ann.id not in read_ids:
            AnnouncementRead.objects.get_or_create(announcement=ann, user=request.user)
 
    # Count unread announcements for bell
    unread_announcements_count = sum(1 for ann in announcements if ann.id not in read_ids)
 
    # ----------------------------
    # CONTEXT
    # ----------------------------
    context = {
        "students": students,
        "student": student,
        "total_present": total_present,
        "total_absent": total_absent,
        "term_results": term_results,
        "session_results": session_results,
        "teacher_remarks": teacher_remarks,
        "subject_teacher_remarks": subject_teacher_remarks,
        "upcoming_exams": upcoming_exams,
        "next_exam": next_exam,
        "announcements": announcements,
        "unread_announcements_count": unread_announcements_count,
        "now": now,
        "result_periods": parent_result_periods,
        "active_session": active_session,
        "active_term": active_term,
    }
 
    return render(request, "parents/dashboard.html", context)





@csrf_exempt
def user_logout(request):
    logout(request)
    return redirect('users:home')




# ------------------------
# Staff Management
# ------------------------


@login_required
def create_staff(request):
    can_create = (
        request.user.is_superuser or
        (request.user.staff is not None and request.user.staff.can_create_staff)
    )

    if not can_create:
        messages.error(request, "You do not have permission to create staff.")
        return redirect('dashboard:router')

    if request.method == 'POST':
        user_form = UserForm(request.POST)
        staff_form = StaffForm(request.POST)

        if user_form.is_valid() and staff_form.is_valid():
            with transaction.atomic():
                user = user_form.save(commit=False)
                user.set_password(user_form.cleaned_data['password'])
                user.is_staff = True
                user.is_staff_user = True
                user.tenant = request.tenant  # scope user to this school
                user.save()

                staff = staff_form.save(commit=False)
                staff.user = user
                staff.save()

            messages.success(request, f"Staff {user.username} created successfully!")
            return redirect('dashboard:router')
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        user_form = UserForm()
        staff_form = StaffForm()

    return render(
        request,
        'admin/create_staff.html',
        {'user_form': user_form, 'staff_form': staff_form}
    )


@login_required
def register_student(request):
    # Permission check
    can_register = request.user.is_superuser or (request.user.staff is not None and request.user.staff.role == "ADMIN")
    if not can_register:
        messages.error(request, "You do not have permission to register a student.")
        return redirect('dashboard:router')

    if request.method == "POST":
        # If the POST contains a base64 camera snapshot (no file uploaded),
        # decode it into a Django InMemoryUploadedFile so the form handles it normally.
        files = request.FILES.copy()
        b64 = request.POST.get('captured_passport_data', '').strip()
        if b64 and 'passport' not in request.FILES:
            import base64, io
            from django.core.files.uploadedfile import InMemoryUploadedFile
            try:
                header, data = b64.split(',', 1)
                img_bytes = base64.b64decode(data)
                img_file  = InMemoryUploadedFile(
                    file=io.BytesIO(img_bytes),
                    field_name='passport',
                    name='passport_photo.jpg',
                    content_type='image/jpeg',
                    size=len(img_bytes),
                    charset=None,
                )
                files['passport'] = img_file
            except Exception:
                pass  # ignore bad data — form will just have no passport

        form = StudentForm(request.POST, files)
        if form.is_valid():
            admission_no = form.cleaned_data.get('admission_number') or f"STU{Student.objects.count()+1:04d}"
            student_password = get_random_string(8)
            full_name = form.cleaned_data.get('full_name')
            first_name, last_name = (full_name.strip().split(' ', 1) + [""])[:2]

            # --- Create student user — scoped to this tenant ---
            student_user = User.objects.create_user(
                username=admission_no,
                password=student_password,
                first_name=first_name,
                last_name=last_name,
                is_student=True,
                is_staff_user=False,
                tenant=request.tenant,   # ← scope to this school
            )
            student_user.term_password = student_password
            student_user.save()

            # --- Save student record ---
            student = form.save(commit=False)
            student.user = student_user
            student.tenant = request.tenant   # ← scope to this school
            student.admission_number = admission_no
            student.parent_name = form.cleaned_data.get('parent_name')
            student.parent_email = form.cleaned_data.get('parent_email')
            student.parent_phone = form.cleaned_data.get('parent_phone')

            # --- Create or get parent user ---
            parent_user = None
            parent_password = None
            if student.parent_email:
                parent_user, created = User.objects.get_or_create(
                    username=student.parent_email,
                    tenant=request.tenant,   # ← parent also scoped to this school
                    defaults={
                        "email": student.parent_email,
                        "is_student": False,
                        "is_parent": True,
                        "is_staff_user": False,
                        "tenant": request.tenant,
                        "first_name": student.parent_name.split(" ")[0] if student.parent_name else "",
                        "last_name": " ".join(student.parent_name.split(" ")[1:]) if student.parent_name else "",
                    }
                )

                if created:
                    parent_password = get_random_string(8)
                    parent_user.set_password(parent_password)
                    parent_user.save()
                else:
                    parent_password = None

                student.parent_user = parent_user
                student.parent_term_password = parent_password

            student.save()

            messages.success(request, f"Student {admission_no} registered successfully!")
            return render(request, "students/student_success.html", {
                "student": student,
                "student_password": student_password,
                "parent_password": parent_password
            })

    else:
        form = StudentForm()

    return render(request, "students/register_student.html", {"form": form})






@login_required
def edit_student(request, student_id):
    """
    Edit student profile and information
    """
    # Permission check
    can_edit = request.user.is_superuser or (
        request.user.staff is not None and request.user.staff.role == "ADMIN"
    )
    
    if not can_edit:
        messages.error(request, "You do not have permission to edit student profiles.")
        return redirect('dashboard:router')
    
    student = get_object_or_404(Student, id=student_id)
    
    if request.method == "POST":
        # Handle base64 captured passport (camera capture)
        import base64, uuid
        from django.core.files.base import ContentFile
        mutable_files = request.FILES.copy() if request.FILES else {}
        b64 = request.POST.get('captured_passport_data', '').strip()
        if b64 and 'passport' not in request.FILES:
            try:
                fmt, imgstr = b64.split(';base64,')
                img_data = base64.b64decode(imgstr)
                fname = f"passport_{uuid.uuid4().hex[:8]}.jpg"
                from django.core.files.uploadedfile import InMemoryUploadedFile
                import io
                img_file = InMemoryUploadedFile(
                    io.BytesIO(img_data), 'passport', fname,
                    'image/jpeg', len(img_data), None
                )
                mutable_files['passport'] = img_file
            except Exception:
                pass
        files_data = mutable_files if mutable_files else request.FILES
        form = StudentForm(request.POST, files_data, instance=student)
        
        if form.is_valid():
            with transaction.atomic():
                # Update student record
                student = form.save(commit=False)
                
                # Update full name in user model
                full_name = form.cleaned_data.get('full_name')
                if full_name:
                    first_name, last_name = (full_name.strip().split(' ', 1) + [""])[:2]
                    student.user.first_name = first_name
                    student.user.last_name = last_name
                    student.user.save()
                
                # Update parent information
                student.parent_name = form.cleaned_data.get('parent_name')
                student.parent_email = form.cleaned_data.get('parent_email')
                student.parent_phone = form.cleaned_data.get('parent_phone')
                
                # Handle parent user account
                if student.parent_email:
                    parent_user, created = User.objects.get_or_create(
                        username=student.parent_email,
                        tenant=request.tenant,
                        defaults={
                            "email": student.parent_email,
                            "is_student": False,
                            "is_parent": True,
                            "is_staff_user": False,
                            "tenant": request.tenant,
                            "first_name": student.parent_name.split(" ")[0] if student.parent_name else "",
                            "last_name": " ".join(student.parent_name.split(" ")[1:]) if student.parent_name else "",
                        }
                    )
                    
                    # If parent user was just created, generate password
                    if created:
                        parent_password = get_random_string(8)
                        parent_user.set_password(parent_password)
                        parent_user.save()
                        student.parent_term_password = parent_password
                    else:
                        # Update existing parent user info
                        parent_user.email = student.parent_email
                        if student.parent_name:
                            parent_user.first_name = student.parent_name.split(" ")[0]
                            parent_user.last_name = " ".join(student.parent_name.split(" ")[1:])
                        parent_user.save()
                    
                    student.parent_user = parent_user
                
                student.save()
                
                messages.success(
                    request,
                    f"Student profile for {student.full_name} updated successfully!"
                )
                return redirect('users:all_students')
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = StudentForm(instance=student)
    
    context = {
        'form': form,
        'student': student,
    }
    
    return render(request, 'students/edit_student.html', context)


@login_required
def delete_student(request, student_id):
    """
    Delete a student
    """
    # Permission check
    can_delete = request.user.is_superuser or (
        request.user.staff is not None and request.user.staff.role == "ADMIN"
    )
    
    if not can_delete:
        messages.error(request, "You do not have permission to delete students.")
        return redirect('dashboard:router')
    
    student = get_object_or_404(Student, id=student_id)
    student_name = student.full_name
    admission_no = student.admission_number
    
    # Delete student user account
    user = student.user
    student.delete()
    user.delete()
    
    messages.success(
        request,
        f"Student {student_name} ({admission_no}) has been deleted successfully."
    )
    
    return redirect('users:all_students')


@login_required
def reset_student_password(request, student_id):
    """
    Reset student password
    """
    can_reset = request.user.is_superuser or (
        request.user.staff is not None and request.user.staff.role == "ADMIN"
    )
    
    if not can_reset:
        messages.error(request, "You do not have permission to reset passwords.")
        return redirect('dashboard:router')
    
    student = get_object_or_404(Student, id=student_id)
    
    # Generate new password
    new_password = get_random_string(8)
    student.user.set_password(new_password)
    student.user.term_password = new_password
    student.user.save()
    
    messages.success(
        request,
        f"Password reset for {student.full_name}. New password: {new_password}"
    )
    
    return redirect('users:all_students')





# Generate Student & Parent Passwords
# ------------------------
@login_required
def generate_student_password_view(request):
    students = Student.objects.all()

    if request.method == "POST":
        student = get_object_or_404(Student, id=request.POST.get("student_id"))

        # Generate new password for student
        new_student_pass = get_random_string(8)
        student.user.set_password(new_student_pass)
        student.user.term_password = new_student_pass
        student.user.save()

        # Generate new password for parent if email exists
        new_parent_pass = None
        if student.parent_email:
            new_parent_pass = get_random_string(8)
            student.parent_term_password = new_parent_pass
            student.save()

            parent_username = f"parent_{student.id}"
            parent_user, created = User.objects.get_or_create(
                username=parent_username,
                defaults={
                    "email": student.parent_email,
                    "is_student": False,
                    "is_parent": True,
                    "is_staff_user": False,
                    "tenant": student.tenant,
                }
            )
            # Update password for existing parent user
            parent_user.set_password(new_parent_pass)
            parent_user.save()

        # Success message
        msg = f"New password for student {student.user.username}: {new_student_pass}"
        if new_parent_pass:
            msg += f" | New password for parent ({student.parent_name}): {new_parent_pass}"
        messages.success(request, msg)

        return redirect('users:generate_student_password')

    return render(request, "students/generate_password.html", {"students": students})



@login_required
def all_students(request):
    if not (request.user.is_staff_user or request.user.is_superuser):
        messages.error(request, "Unauthorized access")
        return redirect('dashboard:router')

    query            = request.GET.get('q', '').strip()
    selected_class_id = request.GET.get('class_id', '').strip()

    # Build all_classes with student counts for the stat cards + dropdown
    all_classes = Class.objects.annotate(
        student_count=Count('student')
    ).order_by('name')

    # Base queryset
    students = Student.objects.select_related('user', 'class_assigned').all()

    # Class filter
    selected_class = None
    if selected_class_id:
        try:
            selected_class = Class.objects.get(id=selected_class_id)
            students = students.filter(class_assigned_id=selected_class_id)
        except Class.DoesNotExist:
            selected_class_id = ''

    # Search filter
    if query:
        students = students.filter(
            Q(full_name__icontains=query) |
            Q(admission_number__icontains=query) |
            Q(parent_name__icontains=query) |
            Q(parent_email__icontains=query) |
            Q(parent_phone__icontains=query)
        )

    # Parent password placeholder
    for student in students:
        if not student.parent_term_password:
            student.parent_term_password = "(existing parent account)"

    total_count = students.count()

    # Pagination — preserve filters across pages
    paginator = Paginator(students, 25)
    page_number = request.GET.get('page')
    students_page = paginator.get_page(page_number)

    context = {
        'students':          students_page,
        'query':             query,
        'paginator':         paginator,
        'all_classes':       all_classes,
        'selected_class_id': selected_class_id,
        'selected_class':    selected_class,
        'total_count':       total_count,
    }

    return render(request, 'students/all_students.html', context)




@login_required
def export_students_excel(request):
    """Export all students (respecting class/search filters) as .xlsx"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    if not (request.user.is_staff_user or request.user.is_superuser):
        messages.error(request, "Unauthorized access")
        return redirect('dashboard:router')

    query             = request.GET.get('q', '').strip()
    selected_class_id = request.GET.get('class_id', '').strip()

    students = Student.objects.select_related('user', 'class_assigned').all()

    if selected_class_id:
        try:
            students = students.filter(class_assigned_id=selected_class_id)
        except (ValueError, Class.DoesNotExist):
            pass

    if query:
        students = students.filter(
            Q(full_name__icontains=query) |
            Q(admission_number__icontains=query) |
            Q(parent_name__icontains=query) |
            Q(parent_email__icontains=query) |
            Q(parent_phone__icontains=query)
        )

    students = students.order_by('class_assigned__name', 'full_name')

    # ── Build workbook ─────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Students"

    # ── Styles ─────────────────────────────────────────────────────────────
    header_fill   = PatternFill("solid", start_color="1E3A8A")
    subhead_fill  = PatternFill("solid", start_color="3B82F6")
    alt_fill      = PatternFill("solid", start_color="EFF6FF")
    white_fill    = PatternFill("solid", start_color="FFFFFF")
    active_fill   = PatternFill("solid", start_color="DCFCE7")
    inactive_fill = PatternFill("solid", start_color="FEF9C3")

    header_font   = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    subhead_font  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    title_font    = Font(name="Arial", bold=True, color="1E3A8A", size=14)
    body_font     = Font(name="Arial", size=10)
    bold_font     = Font(name="Arial", bold=True, size=10)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── School title row ───────────────────────────────────────────────────
    ws.merge_cells("A1:J1")
    ws["A1"] = "WHITE DIAMONDS ACADEMY — STUDENT EXPORT"
    ws["A1"].font = title_font
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 28

    # Sub-info row
    from django.utils import timezone as tz
    ws.merge_cells("A2:J2")
    label = f"Generated: {tz.localtime().strftime('%d %B %Y  %H:%M')}"
    if query:
        label += f'  |  Search: "{query}"'
    if selected_class_id:
        try:
            cls = Class.objects.get(id=selected_class_id)
            label += f"  |  Class: {cls.name}"
        except Class.DoesNotExist:
            pass
    ws["A2"] = label
    ws["A2"].font = Font(name="Arial", italic=True, color="475569", size=9)
    ws["A2"].alignment = center
    ws.row_dimensions[2].height = 18

    ws.row_dimensions[3].height = 8  # spacer

    # ── Column headers ─────────────────────────────────────────────────────
    headers = [
        "S/N", "Admission No.", "Full Name", "Class",
        "Parent Name", "Parent Email", "Parent Phone",
        "Student Password", "Parent Password", "Status"
    ]
    col_widths = [6, 16, 26, 16, 24, 28, 16, 20, 20, 12]

    for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font     = subhead_font
        cell.fill     = subhead_fill
        cell.alignment = center
        cell.border   = border
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[4].height = 22

    # ── Data rows ──────────────────────────────────────────────────────────
    for idx, student in enumerate(students, start=1):
        row = 4 + idx
        fill = alt_fill if idx % 2 == 0 else white_fill
        status = student.status if hasattr(student, 'status') else 'ACTIVE'
        if status == 'ACTIVE':
            status_fill = active_fill
            status_val  = "Active"
        else:
            status_fill = inactive_fill
            status_val  = "Inactive"

        student_password = ""
        if student.user and hasattr(student.user, 'term_password'):
            student_password = student.user.term_password or ""
        parent_password = student.parent_term_password or "(existing account)"

        row_data = [
            idx,
            student.admission_number or "",
            student.full_name or "",
            student.class_assigned.name if student.class_assigned else "Unassigned",
            student.parent_name or "",
            student.parent_email or "",
            student.parent_phone or "",
            student_password,
            parent_password,
            status_val,
        ]

        for col, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font      = bold_font if col in (1, 2) else body_font
            cell.alignment = center if col in (1, 4, 8, 9, 10) else left
            cell.border    = border
            if col == 10:
                cell.fill = status_fill
            else:
                cell.fill = fill
        ws.row_dimensions[row].height = 18

    # ── Summary row ────────────────────────────────────────────────────────
    summary_row = 4 + students.count() + 1
    ws.merge_cells(f"A{summary_row}:C{summary_row}")
    ws[f"A{summary_row}"] = f"Total: {students.count()} student(s)"
    ws[f"A{summary_row}"].font   = Font(name="Arial", bold=True, color="1E3A8A", size=10)
    ws[f"A{summary_row}"].fill   = header_fill
    ws[f"A{summary_row}"].alignment = left
    ws.row_dimensions[summary_row].height = 20

    # Freeze panes below header
    ws.freeze_panes = "A5"

    # ── Stream response ────────────────────────────────────────────────────
    import io
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = "students_export.xlsx"
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@login_required
def promote_students(request):
    if not (request.user.is_staff_user or request.user.is_superuser):
        messages.error(request, "Unauthorized access.")
        return redirect('dashboard:router')

    active_session = get_active_session()
    active_term    = get_active_term()

    # Admins and superusers can override the lock via ?force=1
    can_force = request.user.is_superuser or (
        request.user.staff is not None and request.user.staff.role == 'ADMIN'
    )
    force_override  = can_force and request.GET.get('force') == '1'
    promotion_locked = bool(active_session and active_term) and not force_override

    students = Student.objects.select_related("class_assigned").order_by(
        "class_assigned__name", "full_name"
    )
    classes = Class.objects.all().order_by("name")

    filter_class  = request.GET.get('class_assigned', '').strip()
    filter_term   = request.GET.get('current_term', '').strip()
    filter_status = request.GET.get('status', '').strip()
    query         = request.GET.get('q', '').strip()

    if filter_class:
        students = students.filter(class_assigned_id=filter_class)
    if filter_term:
        students = students.filter(current_term=filter_term)
    if filter_status:
        students = students.filter(status=filter_status)
    if query:
        students = students.filter(
            Q(full_name__icontains=query) |
            Q(admission_number__icontains=query)
        )

    if request.method == "POST":
        if promotion_locked:
            messages.error(
                request,
                "Promotion is locked while a term is active. "
                "Use the 'Force Promote' button if you need to proceed."
            )
            return redirect('users:promote_students')

        student_ids  = [sid for sid in request.POST.getlist('student_ids') if sid]
        new_class_id = request.POST.get("new_class", "").strip()
        new_term     = request.POST.get("new_term", "").strip()

        if not student_ids:
            messages.error(request, "Please select at least one student.")
            return redirect('users:promote_students')
        if not new_class_id:
            messages.error(request, "Please select a destination class.")
            return redirect('users:promote_students')
        if not new_term:
            messages.error(request, "Please select a destination term.")
            return redirect('users:promote_students')

        try:
            new_class = Class.objects.get(id=new_class_id)
        except Class.DoesNotExist:
            messages.error(request, "Selected class does not exist.")
            return redirect('users:promote_students')

        updated = Student.objects.filter(id__in=student_ids).update(
            class_assigned=new_class,
            current_term=new_term,
        )

        messages.success(
            request,
            f"{updated} student(s) promoted to {new_class.name} — {new_term}"
            + (f" (Session: {active_session.name})" if active_session else "")
            + (" [Override]" if force_override else "")
        )
        return redirect('users:promote_students')

    context = {
        "students":        students,
        "classes":         classes,
        "terms":           TERM_OPTIONS,
        "statuses":        STATUS_OPTIONS,
        "filter_class":    filter_class,
        "filter_term":     filter_term,
        "filter_status":   filter_status,
        "query":           query,
        "active_session":  active_session,
        "active_term":     active_term,
        "promotion_locked": promotion_locked,
        "can_force":       can_force,
        "force_override":  force_override,
        "student_count":   students.count(),
    }

    return render(request, "students/promote_students.html", context)



# ------------------------
# Class Management
# ------------------------
@login_required
def manage_classes(request):
    if not (request.user.is_staff_user or request.user.is_superuser):
        messages.error(request, "Unauthorized access.")
        return redirect('dashboard:router')

    classes = Class.objects.all()
    form = ClassForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        grade_name = form.cleaned_data['name']
        if Class.objects.filter(name=grade_name).exists():
            messages.warning(request, f"{grade_name} already exists.")
        else:
            form.save()
            messages.success(request, f"{grade_name} created successfully!")
        return redirect('users:manage_classes')

    return render(request, 'admin/manage_classes.html', {'classes': classes, 'form': form})


@login_required
def delete_class(request, class_id):
    if not (request.user.is_staff_user or request.user.is_superuser):
        messages.error(request, "Unauthorized access.")
        return redirect('dashboard:router')
    cls = get_object_or_404(Class, id=class_id)
    cls_name = cls.name
    cls.delete()
    messages.success(request, f"{cls_name} has been deleted.")
    return redirect('users:manage_classes')


# ------------------------
# Subject Assignment
# ------------------------
@login_required
def assign_subjects(request):
    classes = Class.objects.all()
    subjects = Subject.objects.all()
    selected_class_id = request.GET.get('class')

    selected_class = (
        Class.objects.filter(id=selected_class_id).first()
        if selected_class_id else None
    )

    students = (
        Student.objects.filter(class_assigned=selected_class)
        if selected_class else []
    )

    if request.method == "POST":

        # -------- Assign subjects to CLASS --------
        if 'assign_class_subject' in request.POST:
            class_obj = get_object_or_404(
                Class, id=request.POST.get('school_class')
            )

            for sid in request.POST.getlist('subjects'):
                ClassSubject.objects.get_or_create(
                    school_class=class_obj,
                    subject_id=sid
                )

            messages.success(request, "Subjects assigned to class successfully!")
            return redirect(f"{request.path}?class={class_obj.id}")

        # -------- Assign subjects to STUDENT --------
        if 'assign_student_subject' in request.POST:
            student = get_object_or_404(
                Student, id=request.POST.get('student')
            )

            for sid in request.POST.getlist('subjects'):
                StudentSubject.objects.get_or_create(
                    student=student,
                    subject_id=sid
                )

            messages.success(request, "Subjects assigned to student successfully!")
            return redirect(f"{request.path}?class={selected_class_id}")

    # -------- NEW PART: send assigned subjects to template --------
    assigned_class_subjects = (
        ClassSubject.objects.filter(school_class=selected_class)
        if selected_class else []
    )

    assigned_student_subjects = {
        student.id: StudentSubject.objects.filter(student=student)
        for student in students
    } if selected_class else {}

    context = {
        'classes': classes,
        'subjects': subjects,
        'students': students,
        'selected_class': selected_class,
        'assigned_class_subjects': assigned_class_subjects,
        'assigned_student_subjects': assigned_student_subjects,
    }

    return render(request, 'admin/assign_subject.html', context)




@login_required
def remove_class_subject(request, class_id, subject_id):
    class_subject = get_object_or_404(
        ClassSubject, 
        school_class_id=class_id, 
        subject_id=subject_id
    )
    class_subject.delete()
    messages.success(request, "Subject removed from class successfully!")
    return redirect(f"/subjects/assign/?class={class_id}")


def remove_student_subject(request, student_id, subject_id):
    # 1) Remove the assignment
    StudentSubject.objects.filter(
        student_id=student_id,
        subject_id=subject_id
    ).delete()

    # 2) ALSO remove all related TermResults for that subject
    TermResult.objects.filter(
        student_id=student_id,
        subject_id=subject_id
    ).delete()

    messages.success(request, "Subject removed from student and results cleared.")
    return redirect(request.META.get('HTTP_REFERER', '/'))





# -------------------------------------------------
# Attendance Home (Choose Mark or View)
# -------------------------------------------------
@login_required
def attendance_home(request):
    """
    Landing page after clicking Attendance from sidebar.
    User chooses to Mark or View attendance.
    """
    # 🔐 PERMISSION CHECK
    staff = getattr(request.user, 'staff', None)
    can_mark = (
        request.user.is_superuser or
        (staff and staff.role == 'ADMIN') or
        (staff and getattr(staff, 'can_mark_attendance', False))
    )
    if not can_mark:
        messages.error(request, "You do not have permission to access attendance.")
        return redirect('dashboard:router')
    return render(request, "attendance/home.html")


# -------------------------------------------------
# Select Class → Mark Attendance
# -------------------------------------------------
@login_required
def select_class_mark(request):
    # 🔐 PERMISSION CHECK
    staff = getattr(request.user, 'staff', None)
    can_mark = (
        request.user.is_superuser or
        (staff and staff.role == 'ADMIN') or
        (staff and getattr(staff, 'can_mark_attendance', False))
    )
    if not can_mark:
        messages.error(request, "You do not have permission to mark attendance.")
        return redirect('dashboard:router')

    classes = Class.objects.all()

    if request.method == "POST":
        class_id = request.POST.get("class_id")
        return redirect("users:mark_attendance", class_id=class_id)

    return render(request, "attendance/select_class_mark.html", {
        "classes": classes
    })


# -------------------------------------------------
# Select Class → View Attendance
# -------------------------------------------------
@login_required
def select_class_view(request):
    classes = Class.objects.all()

    if request.method == "POST":
        class_id = request.POST.get("class_id")
        return redirect("users:view_attendance", class_id=class_id)

    return render(request, "attendance/select_class_view.html", {
        "classes": classes
    })


# -------------------------------------------------
# Mark Attendance (Class + Date)
# -------------------------------------------------
@login_required
def mark_attendance(request, class_id):
    # 🔐 PERMISSION CHECK
    _staff = getattr(request.user, 'staff', None)
    _can_mark = (
        request.user.is_superuser or
        (_staff and _staff.role == 'ADMIN') or
        (_staff and getattr(_staff, 'can_mark_attendance', False))
    )
    if not _can_mark:
        messages.error(request, "You do not have permission to mark attendance.")
        return redirect('dashboard:router')

    school_class = get_object_or_404(Class, id=class_id)
    students = Student.objects.filter(class_assigned=school_class)
    date_today = timezone.now().date()

    if request.method == "POST":
        selected_date = request.POST.get("date") or date_today

        for student in students:
            status = request.POST.get(f"status_{student.id}", "A")  # A = Absent
            Attendance.objects.update_or_create(
                student=student,
                school_class=school_class,
                date=selected_date,
                defaults={"status": status}
            )

        messages.success(
            request,
            f"Attendance for {school_class.name} on {selected_date} saved successfully."
        )
        return redirect(
        reverse("users:view_attendance", kwargs={"class_id": school_class.id})
        + f"?date={selected_date}"
        )

    return render(request, "attendance/mark.html", {
        "school_class": school_class,
        "students": students,
        "date_today": date_today
    })


# -------------------------------------------------
# View Attendance (Class + Optional Date Filter)
# -------------------------------------------------
@login_required
def view_attendance(request, class_id):
    school_class = get_object_or_404(Class, id=class_id)
    date_filter = request.GET.get("date")

    records = Attendance.objects.filter(school_class=school_class)

    if date_filter:
        records = records.filter(date=date_filter)

    return render(request, "attendance/view.html", {
        "school_class": school_class,
        "records": records,
        "date_filter": date_filter
    })



@login_required
def attendance_summary(request):
    classes = Class.objects.all()
    selected_class_id = request.GET.get('class_id')
    selected_student_id = request.GET.get('student_id')

    selected_class = None
    students = Student.objects.none()  # Default empty queryset
    summary_data = []
    chart_labels = []
    chart_present = []
    chart_absent = []

    if selected_class_id:
        selected_class = get_object_or_404(Class, id=selected_class_id)
        students = Student.objects.filter(class_assigned=selected_class)

        if selected_student_id:
            students = students.filter(id=selected_student_id)

        for student in students:
            total_present = Attendance.objects.filter(student=student, status='P').count()
            total_absent = Attendance.objects.filter(student=student, status='A').count()
            summary_data.append({
                'student': student,
                'present': total_present,
                'absent': total_absent
            })

            # For chart
            chart_labels.append(student.full_name)
            chart_present.append(total_present)
            chart_absent.append(total_absent)

    context = {
        'classes': classes,
        'students': students,
        'summary_data': summary_data,
        'selected_class': selected_class,
        'selected_student_id': selected_student_id,
        'chart_labels': chart_labels,
        'chart_present': chart_present,
        'chart_absent': chart_absent,
    }
    return render(request, 'attendance/summary.html', context)



@login_required
def edit_attendance(request, class_id, date):
    # 🔐 PERMISSION CHECK
    _staff = getattr(request.user, 'staff', None)
    _can_mark = (
        request.user.is_superuser or
        (_staff and _staff.role == 'ADMIN') or
        (_staff and getattr(_staff, 'can_mark_attendance', False))
    )
    if not _can_mark:
        messages.error(request, "You do not have permission to edit attendance.")
        return redirect('dashboard:router')

    school_class = get_object_or_404(Class, id=class_id)
    students = Student.objects.filter(class_assigned=school_class)

    # Ensure every student has an attendance record for the date
    for student in students:
        Attendance.objects.get_or_create(
            student=student,
            school_class=school_class,
            date=date,
            defaults={'status': 'A'}  # default to Absent
        )

    # Now fetch all attendance records for the class and date
    attendance_records = Attendance.objects.filter(school_class=school_class, date=date)

    if request.method == 'POST':
        for record in attendance_records:
            status = request.POST.get(f'status_{record.student.id}', 'A')
            record.status = status
            record.save()
        messages.success(request, f"Attendance updated for {school_class.name} on {date}")
        return redirect('attendance:view_attendance', class_id=class_id)

    context = {
        'school_class': school_class,
        'attendance_records': attendance_records,
        'date_selected': date
    }
    return render(request, 'attendance/edit.html', context)



@login_required
def delete_attendance(request, class_id, student_id, date):
    # 🔐 PERMISSION CHECK
    _staff = getattr(request.user, 'staff', None)
    _can_mark = (
        request.user.is_superuser or
        (_staff and _staff.role == 'ADMIN') or
        (_staff and getattr(_staff, 'can_mark_attendance', False))
    )
    if not _can_mark:
        messages.error(request, "You do not have permission to delete attendance records.")
        return redirect('dashboard:router')

    record = get_object_or_404(Attendance, school_class_id=class_id, student_id=student_id, date=date)
    record.delete()
    messages.success(request, "Attendance record deleted successfully.")
    return redirect('users:view_attendance', class_id=class_id)





@login_required
def manage_staff_permissions(request):
    """
    View and manage staff permissions
    Only accessible by superuser and admins
    """
    # Permission check
    can_manage = request.user.is_superuser or (
        request.user.staff is not None and request.user.staff.role == 'ADMIN'
    )
    
    if not can_manage:
        messages.error(request, "You do not have permission to manage staff.")
        return redirect('dashboard:router')
    
    # Get all staff with their user info
    from users.models import Staff
    staff_members = Staff.objects.select_related('user').all()
    
    # Search functionality
    query = request.GET.get('q')
    if query:
        staff_members = staff_members.filter(
            Q(user__username__icontains=query) |
            Q(user__first_name__icontains=query) |
            Q(user__last_name__icontains=query) |
            Q(user__email__icontains=query)
        )
    
    # Filter by role
    role_filter = request.GET.get('role')
    if role_filter:
        staff_members = staff_members.filter(role=role_filter)
    
    # Handle permission updates
    if request.method == 'POST':
        staff_id = request.POST.get('staff_id')
        staff = get_object_or_404(Staff, id=staff_id)
        
        # Update role
        new_role = request.POST.get('role')
        if new_role and new_role in ('ADMIN', 'TEACHER', 'ACCOUNT', 'LIB'):
            staff.role = new_role

        # ── Update all permission flags ────────────────────────────────────
        # Use actual fields that exist on the Staff model
        p = request.POST
        staff.can_create_staff      = 'can_create_staff'      in p
        staff.can_manage_students   = 'can_manage_students'   in p
        staff.can_generate_password = 'can_generate_password' in p
        # Exams
        staff.can_manage_exams      = 'can_manage_exams'      in p
        staff.can_view_all_exams    = 'can_view_all_exams'    in p
        # Results
        staff.can_enter_scores      = 'can_enter_scores'      in p
        staff.can_view_all_results  = 'can_view_all_results'  in p
        staff.can_publish_results   = 'can_publish_results'   in p
        # Finance
        staff.can_access_finance    = 'can_access_finance'    in p
        staff.can_approve_payments  = 'can_approve_payments'  in p
        # Hostel
        staff.can_access_hostel     = 'can_access_hostel'     in p
        staff.can_manage_boarders   = 'can_manage_boarders'   in p
        # Communications
        staff.can_send_messages     = 'can_send_messages'     in p
        # Attendance
        staff.can_mark_attendance   = 'can_mark_attendance'   in p
        # Inventory
        staff.can_access_inventory  = 'can_access_inventory'  in p

        # Also update user-level flags
        staff.user.restricted = 'restricted' in p
        staff.user.save(update_fields=['restricted'])

        staff.save()
        
        messages.success(
            request, 
            f"Permissions updated for {staff.user.get_full_name() or staff.user.username}"
        )
        return redirect('users:manage_staff_permissions')
    
    # Define role choices manually (since it's not imported from models)
    role_choices = [
        ('ADMIN', 'Administrator'),
        ('TEACHER', 'Teacher'),
    ]
    
    context = {
        'staff_members': staff_members,
        'role_choices': role_choices,
        'query': query,
        'role_filter': role_filter,
    }
    
    return render(request, 'admin/manage_staff_permissions.html', context)


@login_required
def toggle_staff_restriction(request, staff_id):
    """
    Toggle staff account restriction (enable/disable access)
    """
    can_manage = request.user.is_superuser or (
        request.user.staff is not None and request.user.staff.role == 'ADMIN'
    )
    
    if not can_manage:
        messages.error(request, "You do not have permission to manage staff.")
        return redirect('dashboard:router')
    
    from users.models import Staff
    staff = get_object_or_404(Staff, id=staff_id)
    
    # Toggle restriction
    staff.user.restricted = not staff.user.restricted
    staff.user.save()
    
    status = "restricted" if staff.user.restricted else "enabled"
    messages.success(
        request,
        f"Account for {staff.user.get_full_name() or staff.user.username} has been {status}."
    )
    
    return redirect('users:manage_staff_permissions')


@login_required
def delete_staff(request, staff_id):
    """
    Delete a staff member
    """
    can_manage = request.user.is_superuser or (
        request.user.staff is not None and request.user.staff.role == 'ADMIN'
    )
    
    if not can_manage:
        messages.error(request, "You do not have permission to delete staff.")
        return redirect('dashboard:router')
    
    from users.models import Staff
    staff = get_object_or_404(Staff, id=staff_id)
    
    # Don't allow deleting yourself
    if staff.user == request.user:
        messages.error(request, "You cannot delete your own account.")
        return redirect('users:manage_staff_permissions')
    
    staff_name = staff.user.get_full_name() or staff.user.username
    user = staff.user
    
    staff.delete()
    user.delete()
    
    messages.success(request, f"Staff member {staff_name} has been deleted.")
    return redirect('users:manage_staff_permissions')


# ── Student Subject Manager ────────────────────────────────────────────────────

@login_required
def student_subject_manager(request):
    """
    Manage which subjects a student is offering vs not-offering.

    Logic:
    - ClassSubject defines all subjects for a class.
    - StudentSubject records (if any) override: only listed subjects are "offering".
    - If NO StudentSubject records exist for a student, they offer ALL class subjects.
    - Saving: writes StudentSubject records for checked subjects,
      and marks TermResult.is_not_offering=True for unchecked ones
      across all terms/sessions for the active session.
    """
    from results.models import TermResult
    from academics.utils import get_active_session

    active_session = get_active_session()
    all_classes = Class.objects.all().order_by('name')

    sel_class   = request.GET.get('class', '').strip()
    sel_student = request.GET.get('student', '').strip()
    search_q    = request.GET.get('q', '').strip()

    selected_class   = None
    students_qs      = Student.objects.none()
    selected_student = None
    class_subjects   = []   # ClassSubject records for the class
    offering_ids     = set() # subject IDs student IS offering

    if sel_class:
        selected_class = Class.objects.filter(id=sel_class).first()
        if selected_class:
            students_qs = Student.objects.filter(
                class_assigned=selected_class, status='Active'
            ).order_by('full_name')
            if search_q:
                from django.db.models import Q
                students_qs = students_qs.filter(
                    Q(full_name__icontains=search_q) |
                    Q(admission_number__icontains=search_q)
                ).distinct()

    if sel_student:
        selected_student = Student.objects.select_related('class_assigned').filter(
            id=sel_student
        ).first()
        if selected_student:
            class_subjects = list(
                ClassSubject.objects
                .filter(school_class=selected_student.class_assigned)
                .select_related('subject')
                .order_by('subject__name')
            )
            stu_subj_ids = set(
                StudentSubject.objects
                .filter(student=selected_student)
                .values_list('subject_id', flat=True)
            )
            class_subj_ids = {cs.subject_id for cs in class_subjects}
            # If no overrides exist, student offers everything in the class
            offering_ids = stu_subj_ids if stu_subj_ids else class_subj_ids

    # ── POST: save subject choices ────────────────────────────────────────
    if request.method == 'POST':
        post_student_id = request.POST.get('student_id', '').strip()
        checked_ids = set(int(x) for x in request.POST.getlist('offering_subjects') if x.isdigit())

        target = Student.objects.select_related('class_assigned').filter(id=post_student_id).first()
        if not target:
            messages.error(request, 'Student not found.')
            return redirect(f"{request.path}?class={sel_class}&student={post_student_id}")

        all_class_ids = set(
            ClassSubject.objects
            .filter(school_class=target.class_assigned)
            .values_list('subject_id', flat=True)
        )
        not_offering_ids = all_class_ids - checked_ids

        # Rebuild StudentSubject — only if they're a subset (not full class)
        StudentSubject.objects.filter(student=target).delete()
        if checked_ids != all_class_ids:
            # Only write records when it differs from the class default
            for sid in checked_ids & all_class_ids:
                StudentSubject.objects.create(student=target, subject_id=sid)

        # Mark TermResults is_not_offering for current session across all terms
        if active_session:
            try:
                # Mark not-offering subjects — zero scores, ABS grade
                TermResult.objects.filter(
                    student=target,
                    session=active_session,
                    subject_id__in=not_offering_ids
                ).update(
                    is_not_offering=True,
                    ca1_score=0, ca2_score=0, ca3_score=0,
                    essay_score=0, exam_score=0, total_score=0,
                    grade='ABS', remark='Not Offering'
                )

                # Restore offering subjects
                TermResult.objects.filter(
                    student=target,
                    session=active_session,
                    subject_id__in=checked_ids
                ).update(is_not_offering=False)

            except Exception as db_err:
                messages.error(
                    request,
                    f"Database error saving subject preferences: {db_err}. "
                    "Please run: python manage.py migrate"
                )
                return redirect(
                    f"{request.path}?class={target.class_assigned_id}&student={target.id}"
                )

        off_names = list(Subject.objects.filter(id__in=not_offering_ids).values_list('name', flat=True))
        on_names  = list(Subject.objects.filter(id__in=checked_ids & all_class_ids).values_list('name', flat=True))

        parts = []
        if on_names:  parts.append(f"Offering: {', '.join(on_names)}")
        if off_names: parts.append(f"Not offering: {', '.join(off_names)}")
        messages.success(request, f"{target.full_name} — " + " | ".join(parts))

        return redirect(f"{request.path}?class={target.class_assigned_id}&student={target.id}")

    context = {
        'all_classes':      all_classes,
        'students_qs':      students_qs,
        'selected_class':   selected_class,
        'selected_student': selected_student,
        'class_subjects':   class_subjects,
        'offering_ids':     offering_ids,
        'sel_class':        sel_class,
        'sel_student':      sel_student,
        'search_q':         search_q,
    }
    return render(request, 'admin/student_subject_manager.html', context)

# =============================================================================
# Advanced Subject-Teacher-Class Assignment Manager
# =============================================================================
@login_required
def advanced_subject_assignment(request):
    """
    Advanced page for assigning subjects to teachers per class.
    Features: bulk assign, search/filter, drag-style UI, AJAX operations.
    Only accessible by Superuser and Admin.
    """
    from users.models import Staff, StaffSubjectClass

    # Permission check
    is_admin = request.user.is_superuser or (
        request.user.staff is not None and request.user.staff.role == 'ADMIN'
    )
    if not is_admin:
        messages.error(request, "You do not have permission to manage subject assignments.")
        return redirect('dashboard:router')

    # ── AJAX: Remove single assignment ─────────────────────────────────────
    if request.method == 'POST' and request.POST.get('action') == 'remove_assignment':
        ssc_id = request.POST.get('ssc_id')
        try:
            ssc = StaffSubjectClass.objects.get(id=ssc_id)
            name = f"{ssc.staff.full_name} / {ssc.subject.name} / {ssc.school_class.name}"
            ssc.delete()
            return JsonResponse({'status': 'ok', 'message': f'Removed: {name}'})
        except StaffSubjectClass.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Assignment not found.'}, status=404)

    # ── AJAX: Add single assignment ────────────────────────────────────────
    if request.method == 'POST' and request.POST.get('action') == 'add_assignment':
        staff_id   = request.POST.get('staff_id')
        subject_id = request.POST.get('subject_id')
        class_id   = request.POST.get('class_id')
        try:
            staff   = Staff.objects.get(id=staff_id)
            subject = Subject.objects.get(id=subject_id)
            sclass  = Class.objects.get(id=class_id)
            ssc, created = StaffSubjectClass.objects.get_or_create(
                staff=staff, subject=subject, school_class=sclass
            )
            return JsonResponse({
                'status': 'ok' if created else 'exists',
                'message': ('Assigned successfully.' if created else 'Already assigned.'),
                'ssc_id': ssc.id,
                'staff_name': staff.full_name,
                'subject_name': subject.name,
                'class_name': sclass.name,
            })
        except (Staff.DoesNotExist, Subject.DoesNotExist, Class.DoesNotExist) as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

    # ── AJAX: Bulk assign all subjects of a class to a teacher ─────────────
    if request.method == 'POST' and request.POST.get('action') == 'bulk_assign':
        staff_id = request.POST.get('staff_id')
        class_id = request.POST.get('class_id')
        subject_ids = request.POST.getlist('subject_ids')
        try:
            staff  = Staff.objects.get(id=staff_id)
            sclass = Class.objects.get(id=class_id)
            created_count = 0
            for sid in subject_ids:
                subject = Subject.objects.get(id=sid)
                _, created = StaffSubjectClass.objects.get_or_create(
                    staff=staff, subject=subject, school_class=sclass
                )
                if created:
                    created_count += 1
            return JsonResponse({
                'status': 'ok',
                'message': f'{created_count} subject(s) assigned to {staff.full_name}.',
            })
        except (Staff.DoesNotExist, Class.DoesNotExist) as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

    # ── AJAX: Clear all assignments for a teacher in a class ───────────────
    if request.method == 'POST' and request.POST.get('action') == 'clear_teacher_class':
        staff_id = request.POST.get('staff_id')
        class_id = request.POST.get('class_id')
        deleted, _ = StaffSubjectClass.objects.filter(
            staff_id=staff_id, school_class_id=class_id
        ).delete()
        return JsonResponse({'status': 'ok', 'message': f'{deleted} assignment(s) removed.'})

    # ── GET: Build context ─────────────────────────────────────────────────
    teachers = Staff.objects.select_related('user').filter(
        role__in=['TEACHER', 'ADMIN']
    ).order_by('full_name')

    classes  = Class.objects.all().order_by('name')
    subjects = Subject.objects.all().order_by('name')

    # Build assignment map: {class_id: {subject_id: [staff, ...]}}
    all_assignments = StaffSubjectClass.objects.select_related(
        'staff', 'subject', 'school_class'
    ).all()

    # Per-teacher summary for sidebar
    teacher_summary = []
    for t in teachers:
        assigns = all_assignments.filter(staff=t)
        teacher_summary.append({
            'teacher': t,
            'count': assigns.count(),
            'classes': list(assigns.values_list('school_class__name', flat=True).distinct()),
        })

    # Per-class assignment matrix
    class_matrix = []
    for cls in classes:
        cls_assigns = all_assignments.filter(school_class=cls)
        cls_subjects = ClassSubject.objects.filter(school_class=cls).select_related('subject')
        subject_rows = []
        for cs in cls_subjects:
            assigned_teachers = cls_assigns.filter(subject=cs.subject).select_related('staff')
            subject_rows.append({
                'subject': cs.subject,
                'teachers': assigned_teachers,
            })
        class_matrix.append({
            'class': cls,
            'subjects': subject_rows,
            'total_subjects': cls_subjects.count(),
            'covered': cls_assigns.values('subject').distinct().count(),
        })

    context = {
        'teachers':       teachers,
        'classes':        classes,
        'subjects':       subjects,
        'teacher_summary': teacher_summary,
        'class_matrix':   class_matrix,
        'all_assignments': all_assignments,
    }
    return render(request, 'admin/advanced_subject_assignment.html', context)