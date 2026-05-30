"""
finance/views.py — WDA Finance ERP  (thin views, fat service layer)
"""
import json
from decimal import Decimal, InvalidOperation
from datetime import date

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Q, Sum
from django.http import FileResponse, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST

from academics.models import AcademicSession
from academics.utils import get_active_session, get_active_term
from users.models import Student

from finance.models import (
    AuditLog, ChildBilling, FeePayment, FeeStructure,
    Invoice, InvoiceItem, Installment,
    ParentBilling, Payment, Purchase, SchoolItem,
    TopUpRequest, Transaction, Wallet, _gen_ref,
)
from finance.services.payment_service import (
    approve_payment, build_statement, create_payment,
    ensure_child_billing, export_all_billing_excel,
    export_statement_excel, generate_payment_receipt_pdf,
    get_finance_summary, handle_paystack_webhook,
    initialize_paystack_payment, reject_payment,
    verify_paystack_signature,
)


# ── Permission helpers ──────────────────────────────────────────────────────

def _is_finance_staff(user):
    if user.is_superuser: return True
    if getattr(user, "is_staff_user", False): return True
    try: return user.staff.role in ("ADMIN", "ACCOUNT")
    except Exception: return False


def _require_finance_staff(fn):
    def wrap(req, *a, **kw):
        if not req.user.is_authenticated or not _is_finance_staff(req.user):
            return HttpResponse("Finance staff only.", status=403)
        return fn(req, *a, **kw)
    wrap.__name__ = fn.__name__
    return wrap


def _get_student_for_parent(request):
    """Return the primary student linked to this parent login."""
    u = request.user.username
    if u.startswith("parent_"):
        try:
            return Student.objects.select_related("class_assigned").get(id=int(u.split("_")[1]))
        except (ValueError, Student.DoesNotExist):
            return None
    email = request.user.email
    return Student.objects.filter(parent_email=email).first() if email else None


def _get_all_children(request):
    """
    Return QuerySet of ALL students linked to this parent.
    Rule: siblings = students sharing the same parent_email.
    The shared wallet belongs to the oldest sibling (lowest id).
    """
    u = request.user.username
    if u.startswith("parent_"):
        try:
            primary = Student.objects.get(id=int(u.split("_")[1]))
            if primary.parent_email:
                return Student.objects.filter(
                    parent_email=primary.parent_email
                ).select_related("class_assigned").order_by("id")
            return Student.objects.filter(id=primary.id).select_related("class_assigned")
        except (ValueError, Student.DoesNotExist):
            return Student.objects.none()
    email = request.user.email
    if email:
        return Student.objects.filter(parent_email=email).select_related("class_assigned").order_by("id")
    return Student.objects.none()


def _get_family_wallet(children_qs):
    """
    One shared wallet for the whole family = wallet of the oldest sibling.
    All siblings share it so parent doesn't need to fund each child separately.
    """
    first = children_qs.first()
    if not first:
        return None
    return Wallet.get_or_create_for_student(first)


def _get_family_wallet_for_student(student):
    """
    Given ANY student, find their family wallet:
    - All siblings share the wallet owned by the oldest sibling (lowest id).
    - Used by admin views so they never see a blank child wallet.
    """
    if not student:
        return None
    if student.parent_email:
        oldest = Student.objects.filter(
            parent_email=student.parent_email
        ).order_by('id').first()
    else:
        oldest = student
    return Wallet.get_or_create_for_student(oldest)


def _get_siblings(student):
    """Return all siblings (same parent_email) ordered by id (oldest first)."""
    if student.parent_email:
        return Student.objects.filter(
            parent_email=student.parent_email
        ).select_related('class_assigned').order_by('id')
    return Student.objects.filter(id=student.id).select_related('class_assigned')


# ── PARENT: Wallet Dashboard ────────────────────────────────────────────────

@login_required
def parent_wallet(request):
    children = _get_all_children(request)
    if not children.exists():
        # Give a helpful diagnosis
        u = request.user.username
        email = request.user.email or ""
        if u.startswith("parent_"):
            try:
                sid = int(u.split("_")[1])
                from users.models import Student as _St
                if not _St.objects.filter(id=sid).exists():
                    messages.error(request, f"Student account (ID {sid}) not found. Contact admin.")
                else:
                    messages.error(request, "Student found but wallet is not set up yet. Contact admin.")
            except Exception:
                messages.error(request, "Invalid parent account. Contact admin.")
        elif email:
            messages.error(request, f"No student linked to email {email}. Contact admin.")
        else:
            messages.error(request, "Parent account has no email and is not linked to any student. Contact admin.")
        return redirect("users:parent_dashboard")

    # Which child is currently selected (for per-child tabs)
    sel_child_id = request.GET.get("child")
    if sel_child_id:
        student = children.filter(id=sel_child_id).first() or children.first()
    else:
        student = children.first()

    # Shared family wallet (oldest sibling owns it)
    wallet         = _get_family_wallet(children)
    txns           = wallet.transactions.all()[:50]
    topups         = wallet.topup_requests.all()[:10]
    payments       = wallet.payments.all()[:10]
    active_session = get_active_session()
    active_term    = get_active_term()

    # Per-child data for SELECTED child
    invoices  = student.invoices.all()[:5]
    purchases = student.purchases.select_related("item").all()[:10]
    fees      = FeePayment.objects.filter(student=student).select_related("fee_structure")
    cb = pb = None
    if active_session and active_term:
        pb, cb = ensure_child_billing(student, active_term, active_session)

    sc = student.class_assigned
    items = SchoolItem.objects.filter(is_active=True, stock_qty__gt=0).filter(
        Q(school_class=sc) | Q(school_class__isnull=True) |
        Q(assigned_student=student) | Q(assigned_student__isnull=True)
    ).distinct()

    paid_ids    = fees.filter(status="PAID").values_list("fee_structure_id", flat=True)
    unpaid_fees = FeeStructure.objects.filter(is_active=True).filter(
        Q(school_class=sc) | Q(school_class__isnull=True)
    ).exclude(id__in=paid_ids)

    total_spent  = wallet.transactions.filter(
        txn_type=Transaction.DEBIT, status="COMPLETED"
    ).aggregate(s=Sum("amount"))["s"] or Decimal("0")
    total_funded = wallet.transactions.filter(
        txn_type=Transaction.CREDIT, status="COMPLETED"
    ).aggregate(s=Sum("amount"))["s"] or Decimal("0")

    # Summary across ALL children for this term
    all_children_data = []
    for ch in children:
        ch_fees = FeePayment.objects.filter(student=ch).select_related("fee_structure")
        ch_unpaid = FeeStructure.objects.filter(is_active=True).filter(
            Q(school_class=ch.class_assigned) | Q(school_class__isnull=True)
        ).exclude(id__in=ch_fees.filter(status="PAID").values_list("fee_structure_id", flat=True))
        total_owed = sum(f.amount for f in ch_unpaid)
        all_children_data.append({
            "student":   ch,
            "fees":      ch_fees,
            "unpaid":    ch_unpaid,
            "total_owed": total_owed,
            "is_active":  ch.id == student.id,
        })

    # Hostel billing for selected student
    hostel_bill = None
    boarder_profile_data = None
    try:
        from hostel.models import BoarderProfile, HostelTermBilling
        boarder_profile_data = BoarderProfile.objects.filter(student=student).first()
        if boarder_profile_data and active_session and active_term:
            hostel_bill = HostelTermBilling.objects.filter(
                boarder=boarder_profile_data,
                session=active_session,
                term=active_term,
            ).first()
    except Exception:
        pass

    return render(request, "finance/parent_wallet.html", dict(
        student=student,
        children=children,
        all_children_data=all_children_data,
        wallet=wallet, txns=txns, topups=topups,
        payments=payments, invoices=invoices, purchases=purchases,
        fees=fees, unpaid_fees=unpaid_fees, items=items,
        total_spent=total_spent, total_funded=total_funded,
        child_billing=cb, parent_billing=pb,
        boarder_profile=boarder_profile_data,
        hostel_bill=hostel_bill,
    ))



@login_required
@require_POST
def parent_pay_hostel(request):
    """Parent pays their hostel bill from wallet."""
    student = _get_student_for_parent(request)
    if not student:
        messages.error(request, "Account error.")
        return redirect("finance:parent_wallet")
    try:
        from hostel.models import BoarderProfile, HostelTermBilling
        from hostel.services import pay_hostel_bill
    except ImportError:
        messages.error(request, "Hostel module not available.")
        return redirect("finance:parent_wallet")

    billing_id = request.POST.get("billing_id")
    try:
        amount = Decimal(str(request.POST.get("amount", "0")))
    except (InvalidOperation, ValueError):
        messages.error(request, "Invalid amount.")
        return redirect("finance:parent_wallet")

    # Verify the billing belongs to this student's family
    children = _get_all_children(request)
    billing = get_object_or_404(
        HostelTermBilling,
        id=billing_id,
        boarder__student__in=children,
    )
    try:
        pay_hostel_bill(billing, amount, performed_by=request.user)
        messages.success(
            request,
            f"Hostel fee ₦{amount:,.2f} paid for {billing.boarder.student.full_name}."
        )
    except ValueError as e:
        messages.error(request, str(e))

    return redirect("finance:parent_wallet")


@login_required
@require_POST
def submit_payment(request):
    student = _get_student_for_parent(request)
    if not student:
        messages.error(request, "Account error."); return redirect("finance:parent_wallet")
    try:
        amount = Decimal(str(request.POST.get("amount","0")))
        if amount <= 0: raise ValueError("Amount must be positive.")
    except (InvalidOperation, ValueError) as e:
        messages.error(request, str(e)); return redirect("finance:parent_wallet")
    wallet     = Wallet.get_or_create_for_student(student)
    method     = request.POST.get("method","CASH")
    desc       = request.POST.get("description","School fee payment").strip()
    invoice_id = request.POST.get("invoice_id") or None
    invoice = Invoice.objects.filter(id=invoice_id, student=student).first() if invoice_id else None
    pmt = create_payment(wallet=wallet, amount=amount, method=method,
        description=desc, invoice=invoice, created_by=request.user,
        proof=request.FILES.get("proof"))
    messages.success(request, f"Payment ₦{amount:,.2f} submitted (Ref: {pmt.reference}). Awaiting approval.")
    return redirect("finance:parent_wallet")


@login_required
@require_POST
def request_topup(request):
    student = _get_student_for_parent(request)
    if not student:
        return JsonResponse({"ok":False,"error":"Account error"}, status=403)
    try:
        amount = Decimal(str(request.POST.get("amount","0")))
        if amount <= 0: raise ValueError()
    except Exception:
        messages.error(request,"Invalid amount."); return redirect("finance:parent_wallet")
    wallet = Wallet.get_or_create_for_student(student)
    TopUpRequest.objects.create(
        wallet=wallet, amount=amount,
        method=request.POST.get("method","CASH"),
        reference=request.POST.get("reference","").strip(),
        payment_proof=request.FILES.get("payment_proof"), status="PENDING")
    messages.success(request, f"Top-up ₦{amount:,.2f} submitted.")
    return redirect("finance:parent_wallet")


@login_required
@require_POST
def pay_fee(request):
    student = _get_student_for_parent(request)
    if not student:
        messages.error(request,"Account error."); return redirect("finance:parent_wallet")
    fee = get_object_or_404(FeeStructure, id=request.POST.get("fee_id"), is_active=True)

    # Allow paying for a specific child (pass child_id in form)
    children = _get_all_children(request)
    child_id = request.POST.get("child_id")
    if child_id:
        target_student = children.filter(id=child_id).first() or student
    else:
        target_student = student

    # Shared family wallet
    wallet = _get_family_wallet(children)
    if not wallet:
        messages.error(request, "Wallet not found."); return redirect("finance:parent_wallet")

    if FeePayment.objects.filter(student=target_student, fee_structure=fee, status="PAID").exists():
        messages.warning(request, f"Already paid: {fee.name} for {target_student.full_name}.")
        return redirect(f"%(url)s?child={target_student.id}" % {"url": "/finance/wallet/"})
    if wallet.balance < fee.amount:
        messages.error(request, f"Insufficient balance. Need ₦{fee.amount:,.2f}.")
        return redirect("finance:parent_wallet")
    ref = _gen_ref("FEE")
    txn = wallet.debit(amount=fee.amount,
                       description=f"Fee: {fee.name} — {target_student.full_name}",
                       ref=ref, performed_by=request.user, category="FEE")
    fp, _ = FeePayment.objects.get_or_create(student=target_student, fee_structure=fee,
                                              defaults={"balance_due":fee.amount})
    fp.amount_paid = fee.amount; fp.balance_due = Decimal("0")
    fp.status="PAID"; fp.transaction=txn; fp.paid_at=timezone.now(); fp.save()
    messages.success(request, f"₦{fee.amount:,.2f} paid for {fee.name} ({target_student.full_name}). Ref: {ref}")
    return redirect("finance:receipt_txn", txn_id=txn.id)


@login_required
@require_POST
def buy_item(request):
    student = _get_student_for_parent(request)
    if not student:
        messages.error(request,"Account error."); return redirect("finance:parent_wallet")
    item = get_object_or_404(SchoolItem, id=request.POST.get("item_id"), is_active=True)
    qty  = max(1, int(request.POST.get("quantity", 1)))

    # Allow buying for a specific child
    children = _get_all_children(request)
    child_id = request.POST.get("child_id")
    if child_id:
        target_student = children.filter(id=child_id).first() or student
    else:
        target_student = student

    if item.stock_qty < qty:
        messages.error(request, f"Only {item.stock_qty} in stock.")
        return redirect("finance:parent_wallet")

    total  = item.price * qty
    wallet = _get_family_wallet(children)
    if not wallet:
        messages.error(request, "Wallet not found."); return redirect("finance:parent_wallet")

    if wallet.balance < total:
        messages.error(request, f"Insufficient balance. Need ₦{total:,.2f}, have ₦{wallet.balance:,.2f}.")
        return redirect("finance:parent_wallet")

    cat = {"TEXTBOOK":"TEXTBOOK","UNIFORM":"UNIFORM","SUPPLY":"SUPPLY"}.get(item.item_type, "OTHER")
    txn = wallet.debit(
        amount=total,
        description=f"Purchase: {item.name} x{qty} for {target_student.full_name}",
        ref=_gen_ref("PUR"), performed_by=request.user, category=cat,
    )

    # Fix 8: Update inventory stock when item is purchased
    item.stock_qty -= qty
    item.save(update_fields=["stock_qty"])
    if item.inventory_asset:
        from inventory.models import StockMovement
        StockMovement.objects.create(
            asset=item.inventory_asset,
            movement_type="OUT",
            quantity=qty,
            performed_by=request.user,
            reason=f"School shop — {target_student.full_name} ({target_student.admission_number})",
        )

    Purchase.objects.create(
        student=target_student, item=item, quantity=qty,
        unit_price=item.price, total_price=total, transaction=txn,
    )
    messages.success(request, f"Purchased {item.name} x{qty} for {target_student.full_name}. ₦{total:,.2f} deducted.")
    return redirect("finance:receipt_txn", txn_id=txn.id)


@login_required
def transaction_history(request):
    student = _get_student_for_parent(request)
    if not student: return redirect("users:parent_dashboard")
    wallet   = Wallet.get_or_create_for_student(student)
    from_d   = request.GET.get("from")
    to_d     = request.GET.get("to")
    from_dt  = date.fromisoformat(from_d) if from_d else None
    to_dt    = date.fromisoformat(to_d)   if to_d   else None
    rows     = build_statement(wallet, from_dt, to_dt)
    txn_type = request.GET.get("type","")
    category = request.GET.get("category","")
    if txn_type:  rows = [r for r in rows if r["txn"].txn_type == txn_type]
    if category:  rows = [r for r in rows if r["txn"].category  == category]
    if request.GET.get("export") == "excel":
        buf   = export_statement_excel(wallet, from_dt, to_dt)
        fname = f"Statement_{student.admission_number}.xlsx"
        return FileResponse(buf, as_attachment=True, filename=fname,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    return render(request, "finance/transaction_history.html", dict(
        student=student, wallet=wallet, rows=rows, txn_type=txn_type,
        category=category, from_date=from_d or "", to_date=to_d or "",
        categories=Transaction.CATEGORY_CHOICES))


@login_required
def receipt_txn(request, txn_id):
    txn    = get_object_or_404(Transaction, id=txn_id)
    wallet = txn.wallet
    is_owner = (request.user.username == f"parent_{wallet.student_id}"
                or _is_finance_staff(request.user))
    if not is_owner: return HttpResponse("Unauthorized", status=403)
    ctx = {"txn":txn, "student":wallet.student, "wallet":wallet, "today":date.today()}
    if request.GET.get("pdf"):
        from django.template.loader import render_to_string
        try:
            import pdfkit
            html    = render_to_string("finance/receipt_pdf.html", ctx, request=request)
            options = {"enable-local-file-access":None,"quiet":""}
            config  = pdfkit.configuration(wkhtmltopdf="/usr/bin/wkhtmltopdf")
            pdf     = pdfkit.from_string(html, False, options=options, configuration=config)
            resp    = HttpResponse(pdf, content_type="application/pdf")
            resp["Content-Disposition"] = f'attachment; filename="Receipt_{txn.reference}.pdf"'
            return resp
        except Exception as e:
            return HttpResponse(f"<h3>PDF error: {e}</h3>")
    return render(request, "finance/receipt.html", ctx)


@login_required
def receipt_payment(request, payment_id):
    payment = get_object_or_404(Payment, id=payment_id, status=Payment.APPROVED)
    wallet  = payment.wallet
    is_owner = (request.user.username == f"parent_{wallet.student_id}"
                or _is_finance_staff(request.user))
    if not is_owner: return HttpResponse("Unauthorized", status=403)
    if request.GET.get("pdf"):
        try:
            buf  = generate_payment_receipt_pdf(payment)
            resp = HttpResponse(buf.read(), content_type="application/pdf")
            resp["Content-Disposition"] = f'attachment; filename="Receipt_{payment.reference}.pdf"'
            return resp
        except ImportError:
            messages.warning(request, "reportlab required for PDF.")
    return render(request, "finance/receipt_payment.html",
                  {"payment":payment, "student":wallet.student, "today":date.today()})


@login_required
def view_invoice(request, invoice_id):
    student  = _get_student_for_parent(request)
    is_admin = _is_finance_staff(request.user)
    if not student and not is_admin: return HttpResponse("Unauthorized",status=403)
    invoice = get_object_or_404(Invoice, id=invoice_id)
    if student and invoice.student != student and not is_admin:
        return HttpResponse("Unauthorized",status=403)
    return render(request,"finance/invoice_detail.html",
                  {"invoice":invoice,"items":invoice.items.all()})


# ── PAYSTACK ─────────────────────────────────────────────────────────────────

@login_required
@require_POST
def paystack_init(request):
    student = _get_student_for_parent(request)
    if not student:
        messages.error(request,"Account error."); return redirect("finance:parent_wallet")
    try:
        amount = Decimal(str(request.POST.get("amount","0")))
        if amount < 100: raise ValueError("Minimum ₦100.")
    except (InvalidOperation, ValueError) as e:
        messages.error(request,str(e)); return redirect("finance:parent_wallet")
    wallet   = Wallet.get_or_create_for_student(student)
    email    = student.parent_email or request.user.email or "noreply@wda.edu.ng"
    callback = request.build_absolute_uri("/finance/paystack/callback/")
    try:
        auth_url, ref = initialize_paystack_payment(wallet, amount, email, callback)
        request.session["paystack_ref"]    = ref
        request.session["paystack_amount"] = str(amount)
        return redirect(auth_url)
    except ValueError as e:
        messages.error(request,str(e)); return redirect("finance:parent_wallet")


@login_required
def paystack_callback(request):
    ref     = request.GET.get("reference") or request.session.pop("paystack_ref","")
    student = _get_student_for_parent(request)
    if not ref or not student: return redirect("finance:parent_wallet")
    if Payment.objects.filter(paystack_ref=ref).exists():
        messages.info(request,"Payment already recorded."); return redirect("finance:parent_wallet")
    try: amount = Decimal(request.session.pop("paystack_amount","0"))
    except Exception: amount = Decimal("0")
    wallet = Wallet.get_or_create_for_student(student)
    create_payment(wallet=wallet, amount=amount, method="PAYSTACK", paystack_ref=ref,
                   description=f"Paystack {ref}", created_by=request.user)
    messages.success(request, f"Paystack payment recorded (₦{amount:,.2f}). Awaiting approval.")
    return redirect("finance:parent_wallet")


@csrf_exempt
def paystack_webhook(request):
    if request.method != "POST": return HttpResponse(status=405)
    sig = request.headers.get("X-Paystack-Signature","")
    if not verify_paystack_signature(request.body, sig):
        return HttpResponse("Invalid signature", status=400)
    try: payload = json.loads(request.body)
    except json.JSONDecodeError: return HttpResponse("Bad JSON",status=400)
    handle_paystack_webhook(payload)
    return HttpResponse(status=200)


# ── ADMIN: Dashboard ─────────────────────────────────────────────────────────

@login_required
@_require_finance_staff
def admin_finance_dashboard(request):
    active_session = get_active_session(); active_term = get_active_term()
    session_id = request.GET.get("session"); term_id = request.GET.get("term")
    session = AcademicSession.objects.filter(id=session_id).first() if session_id else active_session
    from academics.models import Term as T
    term    = T.objects.filter(id=term_id).first() if term_id else active_term
    summary        = get_finance_summary(session, term)
    pending_topups = TopUpRequest.objects.filter(status="PENDING").select_related("wallet__student")
    pending_pmts   = Payment.objects.filter(status=Payment.PENDING).select_related("wallet__student")
    recent_txns    = Transaction.objects.select_related("wallet__student").order_by("-created_at")[:20]
    low_stock      = SchoolItem.objects.filter(is_active=True, stock_qty__lte=5)
    audit_log      = AuditLog.objects.select_related("performed_by").all()[:20]
    return render(request,"finance/admin_dashboard.html", dict(
        **summary, pending_topups=pending_topups, pending_pmts=pending_pmts,
        recent_txns=recent_txns, low_stock=low_stock, audit_log=audit_log,
        sessions=AcademicSession.objects.all().order_by("-name"),
        terms=T.objects.all().order_by("-id"),
        sel_session=session, sel_term=term))


@login_required
@_require_finance_staff
@require_POST
def review_payment(request, payment_id):
    payment = get_object_or_404(Payment, id=payment_id, status=Payment.PENDING)
    action  = request.POST.get("action"); note = request.POST.get("note","").strip()
    if action == "approve":
        try:
            approve_payment(payment, request.user, request)
            messages.success(request, f"Payment {payment.reference} approved. ₦{payment.amount:,.2f} credited.")
        except ValueError as e:
            messages.error(request, str(e))
    elif action == "reject":
        reject_payment(payment, request.user, note, request)
        messages.warning(request, f"Payment {payment.reference} rejected.")
    return redirect("finance:admin_dashboard")


@login_required
@_require_finance_staff
@require_POST
def review_topup(request, topup_id):
    from django.db import transaction as dbt
    topup  = get_object_or_404(TopUpRequest, id=topup_id, status="PENDING")
    action = request.POST.get("action"); note = request.POST.get("note","").strip()
    if action == "approve":
        with dbt.atomic():
            wallet = topup.wallet
            txn = wallet.credit(amount=topup.amount,
                description=f"Wallet top-up — {topup.get_method_display()}",
                ref=_gen_ref("TOP"), performed_by=request.user, category="TOPUP")
            topup.status="APPROVED"; topup.transaction=txn; topup.note=note
            topup.reviewed_by=request.user; topup.reviewed_at=timezone.now(); topup.save()
        messages.success(request, f"Top-up ₦{topup.amount:,.2f} approved.")
    elif action == "reject":
        topup.status="REJECTED"; topup.note=note
        topup.reviewed_by=request.user; topup.reviewed_at=timezone.now(); topup.save()
        messages.warning(request,"Top-up rejected.")
    return redirect("finance:admin_dashboard")


@login_required
@_require_finance_staff
def admin_student_wallet(request, student_id=None):
    if not student_id or student_id == 0:
        student_id = request.GET.get("student_id") or request.POST.get("student_id")
    if not student_id:
        return render(request,"finance/admin_student_wallet.html",
                      {"student":None,"students_all":Student.objects.all().order_by("full_name")})
    student   = get_object_or_404(Student, id=student_id)
    # Always use the FAMILY wallet (oldest sibling owns it)
    wallet    = _get_family_wallet_for_student(student)
    siblings  = _get_siblings(student)
    txns      = wallet.transactions.all()
    fees      = FeePayment.objects.filter(student=student).select_related("fee_structure")
    purchases = student.purchases.select_related("item").all()
    payments  = wallet.payments.all()
    invoices  = student.invoices.all()
    billings  = student.child_billings.select_related("parent_billing__session","parent_billing__term")

    # All siblings fee summary
    siblings_data = []
    for sib in siblings:
        sib_fees = FeePayment.objects.filter(student=sib).select_related("fee_structure")
        sib_unpaid = FeeStructure.objects.filter(is_active=True).filter(
            Q(school_class=sib.class_assigned) | Q(school_class__isnull=True)
        ).exclude(id__in=sib_fees.filter(status="PAID").values_list("fee_structure_id", flat=True))
        siblings_data.append({
            "student":    sib,
            "fees":       sib_fees,
            "unpaid":     sib_unpaid,
            "total_owed": sum(f.amount for f in sib_unpaid),
            "is_active":  sib.id == student.id,
        })

    total_paid   = fees.filter(status="PAID").aggregate(s=Sum("fee_structure__amount"))["s"] or 0
    total_unpaid = FeeStructure.objects.filter(is_active=True).filter(
        Q(school_class=student.class_assigned) | Q(school_class__isnull=True)
    ).exclude(
        id__in=fees.filter(status="PAID").values_list("fee_structure_id",flat=True)
    ).aggregate(s=Sum("amount"))["s"] or 0
    # Boarder / Hostel context
    try:
        from hostel.models import (
            BoarderProfile, Hostel, Bed,
            HostelTermBilling, CheckInOut,
        )
        boarder_profile  = BoarderProfile.objects.filter(student=student).first()
        hostels_list     = Hostel.objects.filter(is_active=True)
        available_beds   = Bed.objects.filter(
            status='AVAILABLE', room__is_active=True
        ).select_related('room__hostel', 'room__floor').order_by(
            'room__hostel__name', 'room__room_number', 'bed_number'
        )
        hostel_billings  = HostelTermBilling.objects.filter(
            boarder__student=student
        ).select_related('term', 'session').order_by('-session__name') if boarder_profile else []
        active_session_h = get_active_session()
        active_term_h    = get_active_term()
        current_hostel_bill = None
        if boarder_profile and active_session_h and active_term_h:
            current_hostel_bill = HostelTermBilling.objects.filter(
                boarder=boarder_profile, session=active_session_h, term=active_term_h
            ).first()
    except Exception:
        boarder_profile = None; hostels_list = []; available_beds = []
        hostel_billings = []; current_hostel_bill = None

    return render(request,"finance/admin_student_wallet.html", dict(
        student=student, wallet=wallet, txns=txns, fees=fees,
        purchases=purchases, payments=payments, invoices=invoices,
        billings=billings, total_paid=total_paid, total_unpaid=total_unpaid,
        siblings=siblings, siblings_data=siblings_data,
        students_all=Student.objects.all().order_by("full_name"),
        # Hostel / Boarder
        boarder_profile=boarder_profile,
        hostels_list=hostels_list,
        available_beds=available_beds,
        hostel_billings=hostel_billings,
        current_hostel_bill=current_hostel_bill,
    ))



# ─────────────────────────────────────────────────────────────────────────────
# ADMIN — Act on behalf of parent (pay fee / buy item for any student)
# ─────────────────────────────────────────────────────────────────────────────

@login_required
@_require_finance_staff
def admin_pay_on_behalf(request):
    """
    Admin/accountant can pay fees or buy items for any student,
    deducting from the family wallet (no parent login required).
    """
    from django.db.models import Q as Qdb
    if request.method == "GET":
        students = Student.objects.all().order_by("full_name")
        # Get selected student
        sel_id  = request.GET.get("student_id")
        student = None; wallet = None; fees = []; items = []
        paid_fee_ids = []
        if sel_id:
            student  = get_object_or_404(Student, id=sel_id)
            wallet   = _get_family_wallet_for_student(student)
            siblings = _get_siblings(student)
            fee_payments = FeePayment.objects.filter(student=student).select_related("fee_structure")
            paid_fee_ids = list(fee_payments.filter(status="PAID").values_list("fee_structure_id", flat=True))
            sc = student.class_assigned
            fees = FeeStructure.objects.filter(is_active=True).filter(
                Qdb(school_class=sc) | Qdb(school_class__isnull=True)
            ).exclude(id__in=paid_fee_ids)

            # Inventory assets with stock (for issuing textbooks, uniforms etc.)
            from inventory.models import Asset
            assets_with_stock = []
            for a in Asset.objects.filter(status="ACTIVE").select_related("category").order_by("name"):
                stk = a.current_stock()
                if stk > 0:
                    a._stock = stk
                    assets_with_stock.append(a)

            # All siblings summary
            siblings_data = []
            for sib in siblings:
                sib_fees = FeePayment.objects.filter(student=sib).select_related("fee_structure")
                sib_unpaid = FeeStructure.objects.filter(is_active=True).filter(
                    Qdb(school_class=sib.class_assigned) | Qdb(school_class__isnull=True)
                ).exclude(id__in=sib_fees.filter(status="PAID").values_list("fee_structure_id", flat=True))
                siblings_data.append({
                    "student":    sib,
                    "unpaid":     sib_unpaid,
                    "total_owed": sum(f.amount for f in sib_unpaid),
                })

        txns = wallet.transactions.all()[:20] if wallet else []
        return render(request, "finance/admin_pay_on_behalf.html", dict(
            students=students, student=student, wallet=wallet,
            fees=fees if student else [],
            assets=assets_with_stock if student else [],
            siblings=siblings if student else [],
            siblings_data=siblings_data if student else [],
            txns=txns,
        ))

    # POST — perform the action
    action     = request.POST.get("action")  # "pay_fee" or "buy_item"
    student_id = request.POST.get("student_id")
    student    = get_object_or_404(Student, id=student_id)

    wallet = _get_family_wallet_for_student(student)

    if action == "pay_fee":
        fee = get_object_or_404(FeeStructure, id=request.POST.get("fee_id"), is_active=True)
        if FeePayment.objects.filter(student=student, fee_structure=fee, status="PAID").exists():
            messages.warning(request, f"Already paid: {fee.name} for {student.full_name}.")
            return redirect(f"/finance/admin/pay-on-behalf/?student_id={student.id}")
        if wallet.balance < fee.amount:
            messages.error(request, f"Insufficient wallet balance. Need ₦{fee.amount:,.2f}, have ₦{wallet.balance:,.2f}.")
            return redirect(f"/finance/admin/pay-on-behalf/?student_id={student.id}")
        ref = _gen_ref("ABFEE")
        txn = wallet.debit(amount=fee.amount,
                           description=f"[Admin] Fee: {fee.name} — {student.full_name}",
                           ref=ref, performed_by=request.user, category="FEE")
        fp, _ = FeePayment.objects.get_or_create(student=student, fee_structure=fee,
                                                  defaults={"balance_due": fee.amount})
        fp.amount_paid = fee.amount; fp.balance_due = Decimal("0")
        fp.status = "PAID"; fp.transaction = txn; fp.paid_at = timezone.now(); fp.save()
        AuditLog.objects.create(
            performed_by=request.user, action="PAYMENT_APPROVE",
            description=f"Admin paid {fee.name} for {student.full_name} — ₦{fee.amount:,.2f}",
        )
        messages.success(request, f"✓ Paid {fee.name} for {student.full_name}.")
        return redirect("finance:receipt_txn", txn_id=txn.id)

    elif action == "issue_asset":
        from inventory.models import Asset, StockMovement
        asset  = get_object_or_404(Asset, id=request.POST.get("asset_id"), status="ACTIVE")
        qty    = max(1, int(request.POST.get("quantity", 1)))
        try:
            unit_price = Decimal(str(request.POST.get("unit_price", "0")))
        except Exception:
            unit_price = Decimal("0")
        total = unit_price * qty
        if asset.current_stock() < qty:
            messages.error(request, f"Only {asset.current_stock()} units of '{asset.name}' in stock.")
            return redirect(f"/finance/admin/pay-on-behalf/?student_id={student.id}")
        if total > 0 and wallet.balance < total:
            messages.error(request, f"Insufficient balance. Need ₦{total:,.2f}, have ₦{wallet.balance:,.2f}.")
            return redirect(f"/finance/admin/pay-on-behalf/?student_id={student.id}")
        txn = None
        if total > 0:
            txn = wallet.debit(
                amount=total,
                description=f"[Issued] {asset.name} x{qty} — {student.full_name}",
                ref=_gen_ref("ISS"), performed_by=request.user, category="TEXTBOOK",
            )
        StockMovement.objects.create(
            asset=asset, movement_type="OUT", quantity=qty,
            performed_by=request.user,
            reason=f"Issued to {student.full_name} ({student.admission_number})" +
                   (f" — paid ₦{total:,.2f}" if total > 0 else " — free issue"),
        )
        AuditLog.objects.create(
            performed_by=request.user, action="PAYMENT_APPROVE" if total > 0 else "OTHER",
            description=f"Admin issued {asset.name} x{qty} to {student.full_name}" +
                        (f" — ₦{total:,.2f}" if total > 0 else " — no charge"),
        )
        if txn:
            messages.success(request, f"✓ Issued {asset.name} x{qty} to {student.full_name}. ₦{total:,.2f} deducted.")
            return redirect("finance:receipt_txn", txn_id=txn.id)
        else:
            messages.success(request, f"✓ Issued {asset.name} x{qty} to {student.full_name} (no charge).")
            return redirect(f"/finance/admin/pay-on-behalf/?student_id={student.id}")

    messages.error(request, "Unknown action.")
    return redirect("finance:admin_pay_on_behalf")


@login_required
@_require_finance_staff
@require_POST
def admin_adjust_wallet(request):
    student_id  = request.POST.get("student_id")
    action      = request.POST.get("action")
    description = request.POST.get("description","Manual adjustment").strip()
    try:
        amount = Decimal(str(request.POST.get("amount","0")))
        if amount <= 0: raise ValueError()
    except Exception:
        messages.error(request,"Invalid amount."); return redirect("finance:admin_student_wallet")
    student = get_object_or_404(Student, id=student_id)
    wallet  = _get_family_wallet_for_student(student)  # always use family wallet
    try:
        if action == "credit":
            wallet.credit(amount, description, performed_by=request.user)
            messages.success(request, f"₦{amount:,.2f} credited to {student.full_name}.")
        else:
            wallet.debit(amount, description, performed_by=request.user)
            messages.success(request, f"₦{amount:,.2f} debited from {student.full_name}.")
    except ValueError as e:
        messages.error(request, str(e))
    return redirect(f"/finance/admin/wallet/{student.id}/")


@login_required
@_require_finance_staff
def admin_invoices(request):
    if request.method == "POST" and request.POST.get("action") == "generate":
        from academics.models import Term as T
        student = get_object_or_404(Student, id=request.POST.get("student_id"))
        term    = get_object_or_404(T, id=request.POST.get("term_id"))
        sess    = get_object_or_404(AcademicSession, id=request.POST.get("session_id"))
        inv = Invoice.generate_for_student(student, term, sess, created_by=request.user)
        messages.success(request, f"Invoice {inv.invoice_no} created.")
        return redirect("finance:admin_invoices")
    from academics.models import Term as T
    return render(request,"finance/admin_invoices.html", dict(
        invoices=Invoice.objects.select_related("student","term","session").all(),
        students=Student.objects.all().order_by("full_name"),
        terms=T.objects.all().order_by("-id"),
        sessions=AcademicSession.objects.all().order_by("-name")))


@login_required
@_require_finance_staff
def admin_billing_report(request):
    """
    Billing report built directly from FeeStructure + FeePayment.
    Always shows ALL active students — even those who never logged in.
    """
    from academics.models import Term as T
    from django.db.models import Q as Qb, Sum as Sb
    active_session = get_active_session(); active_term = get_active_term()
    session_id = request.GET.get("session"); term_id = request.GET.get("term")
    session = AcademicSession.objects.filter(id=session_id).first() if session_id else active_session
    term    = T.objects.filter(id=term_id).first() if term_id else active_term

    # Class filter (optional)
    class_id = request.GET.get("class")
    from users.models import Class as SchClass
    classes = SchClass.objects.all().order_by("name")

    billing_rows = []
    if session or term:
        fee_qs = FeeStructure.objects.filter(is_active=True).filter(
            Qb(term=term) | Qb(term__isnull=True)
        ).filter(
            Qb(session=session) | Qb(session__isnull=True)
        )

        students_qs = Student.objects.filter(status="Active").select_related("class_assigned")
        if class_id:
            students_qs = students_qs.filter(class_assigned_id=class_id)

        for student in students_qs.order_by("class_assigned__name", "full_name"):
            student_fees = fee_qs.filter(
                Qb(school_class=student.class_assigned) | Qb(school_class__isnull=True)
            )
            total_fee = sum(f.amount for f in student_fees)
            if total_fee == 0:
                continue

            # Get all fee payment records for this student
            fee_payments = FeePayment.objects.filter(
                student=student, fee_structure__in=student_fees
            ).select_related("fee_structure")

            paid_map = {fp.fee_structure_id: fp for fp in fee_payments}
            amount_paid = sum(
                fp.amount_paid for fp in fee_payments
                if fp.status in ("PAID", "PARTIAL", "WAIVED")
            )
            balance = max(Decimal("0"), total_fee - amount_paid)

            # Per-fee breakdown
            fee_details = []
            for fee in student_fees:
                fp = paid_map.get(fee.id)
                fee_details.append({
                    "fee":    fee,
                    "status": fp.status if fp else "UNPAID",
                    "paid":   fp.amount_paid if fp else Decimal("0"),
                })

            if balance > 0:
                overall_status = "PARTIAL" if amount_paid > 0 else "UNPAID"
            else:
                overall_status = "PAID"

            billing_rows.append({
                "student":        student,
                "total_fee":      total_fee,
                "amount_paid":    amount_paid,
                "balance":        balance,
                "status":         overall_status,
                "fee_details":    fee_details,
                "wallet_balance": _get_family_wallet_for_student(student).balance
                                  if student.parent_email or True else Decimal("0"),
            })

    # Excel export
    if request.GET.get("export") == "excel" and billing_rows:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            from io import BytesIO
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Fee Billing"
            ws.merge_cells("A1:H1")
            ws["A1"] = f"WHITE DIAMONDS ACADEMY — Billing Report  {term or ''} / {session or ''}"
            ws["A1"].font = Font(bold=True, size=13, color="064E3B")
            ws["A1"].alignment = Alignment(horizontal="center")
            headers = ["Student","Admission No.","Class","Parent","Total Fee (₦)","Paid (₦)","Balance (₦)","Status"]
            hfill = PatternFill(fill_type="solid", fgColor="1E3A5F")
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col, value=h)
                cell.fill = hfill; cell.font = Font(bold=True, color="FFFFFF", size=10)
                cell.alignment = Alignment(horizontal="center")
            thin = Side(style="thin", color="E2E8F0")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            for ri, row in enumerate(billing_rows, 4):
                s = row["student"]
                ws.cell(row=ri, column=1, value=s.full_name)
                ws.cell(row=ri, column=2, value=s.admission_number)
                ws.cell(row=ri, column=3, value=str(s.class_assigned or ""))
                ws.cell(row=ri, column=4, value=s.parent_name or "")
                ws.cell(row=ri, column=5, value=float(row["total_fee"]))
                ws.cell(row=ri, column=6, value=float(row["amount_paid"]))
                bc = ws.cell(row=ri, column=7, value=float(row["balance"]))
                sc = ws.cell(row=ri, column=8, value=row["status"])
                bc.font = Font(color="DC2626" if row["balance"] > 0 else "16A34A", bold=True)
                sc.font = Font(color="DC2626" if row["status"] == "UNPAID" else
                               "D97706" if row["status"] == "PARTIAL" else "16A34A", bold=True)
                for col in range(1, 9):
                    ws.cell(row=ri, column=col).border = border
                    if ri % 2 == 0:
                        ws.cell(row=ri, column=col).fill = PatternFill(fill_type="solid", fgColor="F0FDF4")
            for i, w in enumerate([28,18,16,22,16,16,16,12], 1):
                ws.column_dimensions[get_column_letter(i)].width = w
            buf = BytesIO(); wb.save(buf); buf.seek(0)
            fname = f"Billing_{term}_{session}.xlsx"
            return FileResponse(buf, as_attachment=True, filename=fname,
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        except Exception as e:
            messages.error(request, f"Export error: {e}")

    return render(request, "finance/admin_billing_report.html", dict(
        billing_rows=billing_rows,
        sel_session=session, sel_term=term, sel_class=class_id,
        classes=classes,
        sessions=AcademicSession.objects.all().order_by("-name"),
        terms=T.objects.all().order_by("-id")))


@login_required
@_require_finance_staff
def manage_fees(request):
    if request.method == "POST":
        action = request.POST.get("action")
        if action == "add":
            try: amount = Decimal(str(request.POST.get("amount","0")))
            except InvalidOperation:
                messages.error(request,"Invalid amount."); return redirect("finance:manage_fees")
            FeeStructure.objects.create(
                name=request.POST.get("name","").strip(),
                description=request.POST.get("description","").strip(),
                amount=amount,
                category=request.POST.get("category","FEE"),
                term_id=request.POST.get("term_id") or None,
                session_id=request.POST.get("session_id") or None,
                school_class_id=request.POST.get("school_class") or None,
                is_compulsory=request.POST.get("is_compulsory") == "on",
                due_date=request.POST.get("due_date") or None)
            messages.success(request,"Fee structure added.")
        elif action == "delete":
            FeeStructure.objects.filter(id=request.POST.get("fee_id")).update(is_active=False)
            messages.success(request,"Fee deactivated.")
        return redirect("finance:manage_fees")
    from academics.models import Term as T
    from users.models import Class
    return render(request,"finance/manage_fees.html", dict(
        fees=FeeStructure.objects.filter(is_active=True).select_related("school_class","term","session"),
        classes=Class.objects.all().order_by("name"),
        terms=T.objects.all().order_by("-id"),
        sessions=AcademicSession.objects.all().order_by("-name"),
        category_choices=Transaction.CATEGORY_CHOICES))


@login_required
@_require_finance_staff
def manage_items(request):
    if request.method == "POST":
        action = request.POST.get("action")
        if action == "add":
            try:
                price = Decimal(str(request.POST.get("price","0")))
                stock = int(request.POST.get("stock_qty",0))
            except Exception:
                messages.error(request,"Invalid."); return redirect("finance:manage_items")
            SchoolItem.objects.create(
                name=request.POST.get("name","").strip(),
                description=request.POST.get("description","").strip(),
                item_type=request.POST.get("item_type","OTHER"),
                price=price, stock_qty=stock,
                school_class_id=request.POST.get("school_class") or None,
                assigned_student_id=request.POST.get("assigned_student") or None,
                inventory_asset_id=request.POST.get("inventory_asset") or None,
                image=request.FILES.get("image"))
            messages.success(request,"Item added.")
        elif action == "restock":
            item = get_object_or_404(SchoolItem, id=request.POST.get("item_id"))
            item.stock_qty += int(request.POST.get("qty",0))
            item.save(update_fields=["stock_qty"])
            messages.success(request,f"Stock updated: {item.name} = {item.stock_qty}")
        elif action == "delete":
            SchoolItem.objects.filter(id=request.POST.get("item_id")).update(is_active=False)
            messages.success(request,"Item removed.")
        return redirect("finance:manage_items")
    from users.models import Class
    from inventory.models import Asset
    return render(request,"finance/manage_items.html", dict(
        items=SchoolItem.objects.filter(is_active=True).select_related("school_class"),
        classes=Class.objects.all().order_by("name"),
        assets=Asset.objects.filter(status="ACTIVE").order_by("name")))


@login_required
@require_POST
def waive_fee(request):
    if not request.user.is_superuser: return HttpResponse("Unauthorized",status=403)
    student = get_object_or_404(Student, id=request.POST.get("student_id"))
    fee     = get_object_or_404(FeeStructure, id=request.POST.get("fee_id"))
    note    = request.POST.get("note","").strip()
    fp, _   = FeePayment.objects.get_or_create(student=student, fee_structure=fee,
                                                defaults={"balance_due":fee.amount})
    fp.status="WAIVED"; fp.note=note; fp.waived_by=request.user; fp.paid_at=timezone.now(); fp.save()
    messages.success(request, f"{fee.name} waived for {student.full_name}.")
    return redirect(f"/finance/admin/wallet/{student.id}/")





# ─────────────────────────────────────────────────────────────────────────────
# FINANCE — Boarder Management from Wallet Page
# ─────────────────────────────────────────────────────────────────────────────

@login_required
@_require_finance_staff
@require_POST
def make_boarder(request, student_id):
    """
    Make a student a boarder (or update their boarding status) from the
    finance wallet page. Assigns a bed, sets type and creates boarder profile.
    """
    student = get_object_or_404(Student, id=student_id)
    try:
        from hostel.models import BoarderProfile, Bed
        from hostel.services import assign_bed
    except ImportError:
        messages.error(request, "Hostel module not available.")
        return redirect(f"/finance/admin/wallet/{student_id}/")

    student_type = request.POST.get('student_type', 'BOARDER')
    bed_id       = request.POST.get('bed_id') or None
    session_id   = request.POST.get('session_id') or None

    profile, created = BoarderProfile.objects.get_or_create(
        student=student,
        defaults={'student_type': student_type}
    )
    profile.student_type = student_type
    # Ensure boarders are ACTIVE status
    if student_type in ('BOARDER', 'WEEKLY'):
        profile.status = 'ACTIVE'
    if session_id:
        profile.session_id = session_id
    save_fields = ['student_type', 'status', 'updated_at']
    if session_id:
        save_fields.append('session_id')
    profile.save(update_fields=save_fields)

    if bed_id:
        bed = get_object_or_404(Bed, id=bed_id)
        try:
            assign_bed(profile, bed, performed_by=request.user)
            messages.success(request,
                f"{student.full_name} set as {profile.get_student_type_display()} "
                f"and assigned to Room {bed.room.room_number}, Bed {bed.bed_number} "
                f"({bed.room.hostel.name}).")
        except ValueError as e:
            messages.error(request, str(e))
    else:
        messages.success(request,
            f"{student.full_name} set as {profile.get_student_type_display()}. "
            "No bed assigned yet.")

    # Auto-generate hostel bill for current term if type is boarder
    if student_type in ('BOARDER', 'WEEKLY'):
        try:
            active_sess = get_active_session()
            active_trm  = get_active_term()
            if active_sess and active_trm:
                from hostel.services import generate_hostel_bills
                # Generate only for this student
                from hostel.models import HostelTermBilling, HostelFeeStructure
                from django.db.models import Q as Qhs
                if not HostelTermBilling.objects.filter(
                    boarder=profile, session=active_sess, term=active_trm
                ).exists():
                    fee_qs = HostelFeeStructure.objects.filter(is_active=True)
                    hostel = profile.hostel
                    if hostel:
                        fee = (fee_qs.filter(hostel=hostel, term=active_trm, session=active_sess).first()
                               or fee_qs.filter(hostel=hostel).first()
                               or fee_qs.filter(hostel__isnull=True).first())
                    else:
                        fee = fee_qs.filter(hostel__isnull=True).first()
                    if fee:
                        total = fee.boarding_fee + fee.meal_fee + fee.laundry_fee + fee.other_fee
                        HostelTermBilling.objects.create(
                            boarder=profile, term=active_trm, session=active_sess,
                            boarding_fee=fee.boarding_fee, meal_fee=fee.meal_fee,
                            laundry_fee=fee.laundry_fee, other_fee=fee.other_fee,
                            total_fee=total,
                        )
                        messages.info(request,
                            f"Hostel bill of ₦{total:,.2f} generated for "
                            f"{student.full_name} ({active_trm} / {active_sess}).")
        except Exception:
            pass  # billing auto-gen failure is non-critical

    return redirect(f"/finance/admin/wallet/{student_id}/")


@login_required
@_require_finance_staff
@require_POST
def remove_boarder(request, student_id):
    """Remove a student from boarding (check them out of their bed)."""
    student = get_object_or_404(Student, id=student_id)
    try:
        from hostel.models import BoarderProfile
        from hostel.services import unassign_bed
    except ImportError:
        messages.error(request, "Hostel module not available.")
        return redirect(f"/finance/admin/wallet/{student_id}/")

    profile = BoarderProfile.objects.filter(student=student).first()
    if not profile:
        messages.warning(request, f"{student.full_name} has no boarder profile.")
        return redirect(f"/finance/admin/wallet/{student_id}/")

    reason = request.POST.get('reason', 'Removed from boarding via finance module')
    try:
        unassign_bed(profile, reason=reason, performed_by=request.user)
        profile.student_type = 'DAY'
        profile.save(update_fields=['student_type'])
        messages.success(request,
            f"{student.full_name} has been checked out and set back to Day Student.")
    except Exception as e:
        messages.error(request, str(e))

    return redirect(f"/finance/admin/wallet/{student_id}/")


@login_required
@_require_finance_staff
@require_POST
def pay_hostel_fee(request, student_id):
    """Pay hostel bill directly from the finance wallet page."""
    student = get_object_or_404(Student, id=student_id)
    try:
        from hostel.models import BoarderProfile, HostelTermBilling
        from hostel.services import pay_hostel_bill
    except ImportError:
        messages.error(request, "Hostel module not available.")
        return redirect(f"/finance/admin/wallet/{student_id}/")

    billing_id = request.POST.get('billing_id')
    try:
        amount = Decimal(str(request.POST.get('amount', '0')))
    except InvalidOperation:
        messages.error(request, "Invalid amount.")
        return redirect(f"/finance/admin/wallet/{student_id}/")

    billing = get_object_or_404(HostelTermBilling, id=billing_id,
                                 boarder__student=student)
    try:
        pay_hostel_bill(billing, amount, performed_by=request.user)
        messages.success(request,
            f"Hostel bill ₦{amount:,.2f} paid for {student.full_name}.")
    except ValueError as e:
        messages.error(request, str(e))

    return redirect(f"/finance/admin/wallet/{student_id}/")

@login_required
@_require_finance_staff
def admin_audit_log(request):
    return render(request,"finance/admin_audit_log.html",
        {"logs":AuditLog.objects.select_related("performed_by").all()})


@login_required
def wallet_balance_api(request):
    student = _get_student_for_parent(request)
    if not student: return JsonResponse({"balance":"0.00"})
    wallet = Wallet.get_or_create_for_student(student)
    return JsonResponse({"balance":str(wallet.balance)})
