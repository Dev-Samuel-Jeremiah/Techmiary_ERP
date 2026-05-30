"""
finance/services/payment_service.py
=====================================
All business logic lives here. Views stay thin.

Payment → Approval → Wallet Credit → Billing Update →
  Child FIFO Split → Installment Record → Receipt → Statement
"""

import hashlib
import hmac
import uuid
from decimal import Decimal
from io import BytesIO

from django.conf import settings
from django.db import transaction as db_transaction
from django.utils import timezone

from finance.models import (
    AuditLog, ChildBilling, ChildPaymentAllocation,
    FeePayment, FeeStructure, Installment, Invoice,
    InvoiceItem, ParentBilling, Payment, Transaction, Wallet,
)
from users.models import Student


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _gen_ref(prefix="TIT"):
    return prefix + uuid.uuid4().hex[:10].upper()


def _get_client_ip(request):
    xff = request.META.get('HTTP_X_FORWARDED_FOR')
    return xff.split(',')[0].strip() if xff else request.META.get('REMOTE_ADDR', '')


def _audit(user, action, description, request=None, model='', obj_id=None):
    AuditLog.objects.create(
        performed_by=user,
        action=action,
        target_model=model,
        target_id=obj_id,
        description=description,
        ip_address=_get_client_ip(request) if request else None,
    )


# ─────────────────────────────────────────────────────────────────────────────
# 1.  BILLING MANAGEMENT
# ─────────────────────────────────────────────────────────────────────────────

def ensure_child_billing(student, term, session):
    """
    Get or create ParentBilling + ChildBilling for this student/term/session.
    Always uses the FAMILY wallet (oldest sibling by parent_email).
    Updates ChildBilling.paid from FeePayment records each call so it stays in sync.
    Returns (parent_billing, child_billing).
    """
    from django.db.models import Q

    # ── Family wallet (oldest sibling owns it) ─────────────────────────────
    if student.parent_email:
        oldest = Student.objects.filter(
            parent_email=student.parent_email
        ).order_by('id').first() or student
    else:
        oldest = student
    wallet = Wallet.get_or_create_for_student(oldest)

    pb, _ = ParentBilling.objects.get_or_create(
        wallet=wallet, session=session, term=term,
    )
    cb, created = ChildBilling.objects.get_or_create(
        student=student, parent_billing=pb,
    )

    # Always (re)compute total_fee and paid so figures stay accurate
    fees = FeeStructure.objects.filter(is_active=True).filter(
        Q(school_class=student.class_assigned) | Q(school_class__isnull=True)
    ).filter(
        Q(term=term) | Q(term__isnull=True)
    ).filter(
        Q(session=session) | Q(session__isnull=True)
    )
    total_fee = sum(f.amount for f in fees)

    # Amount already paid = sum of paid FeePayment amounts for this student
    from django.db.models import Sum as _Sum
    paid_amount = FeePayment.objects.filter(
        student=student,
        fee_structure__in=fees,
        status__in=['PAID', 'PARTIAL'],
    ).aggregate(s=_Sum('amount_paid'))['s'] or Decimal('0')

    cb.total_fee = total_fee
    cb.paid      = paid_amount
    cb.save(update_fields=['total_fee', 'paid'])
    pb.refresh_from_children()

    return pb, cb


# ─────────────────────────────────────────────────────────────────────────────
# 2.  PAYMENT CREATION
# ─────────────────────────────────────────────────────────────────────────────

def create_payment(wallet, amount, method, description='',
                   invoice=None, paystack_ref='', created_by=None,
                   proof=None):
    """
    Creates a Payment in PENDING status.
    No wallet change happens here — only on approval.
    """
    return Payment.objects.create(
        wallet=wallet,
        amount=Decimal(str(amount)),
        reference=_gen_ref("PMT"),
        method=method,
        status=Payment.PENDING,
        invoice=invoice,
        paystack_ref=paystack_ref,
        description=description,
        proof=proof,
        created_by=created_by,
    )


# ─────────────────────────────────────────────────────────────────────────────
# 3.  PAYMENT APPROVAL  (the critical atomic workflow)
# ─────────────────────────────────────────────────────────────────────────────

def approve_payment(payment, approved_by, request=None):
    """
    Full atomic payment approval:
      1. Credit parent wallet
      2. Record Installment
      3. Update Invoice if linked
      4. FIFO split across ChildBilling rows
      5. Update FeePayment records
      6. Audit log
    Returns the wallet Transaction created.
    """
    if payment.status != Payment.PENDING:
        raise ValueError(f"Payment {payment.reference} is already {payment.status}.")

    with db_transaction.atomic():
        wallet = payment.wallet

        # 1 — Credit wallet
        txn = wallet.credit(
            amount=payment.amount,
            description=payment.description or f"Payment {payment.reference}",
            ref=_gen_ref("APR"),
            performed_by=approved_by,
            category='TOPUP',
            related_payment=payment,
        )

        # 2 — Mark payment approved
        payment.status      = Payment.APPROVED
        payment.approved_by = approved_by
        payment.approved_at = timezone.now()
        payment.save(update_fields=['status', 'approved_by', 'approved_at'])

        # 3 — Installment record
        Installment.objects.create(
            payment=payment,
            invoice=payment.invoice,
            wallet=wallet,
            amount=payment.amount,
            description=payment.description or f"Payment {payment.reference}",
        )

        # 4 — Update invoice balance
        if payment.invoice:
            inv = payment.invoice
            inv.amount_paid += payment.amount
            inv.balance_due  = max(Decimal('0'), inv.total - inv.amount_paid)
            inv.status       = 'PAID' if inv.balance_due == 0 else 'PARTIAL'
            inv.save(update_fields=['amount_paid', 'balance_due', 'status', 'updated_at'])

        # 5 — FIFO split across children
        _split_payment_fifo(payment, wallet)

        # 6 — Audit
        _audit(
            approved_by, 'PAYMENT_APPROVE',
            f"Approved payment {payment.reference} ₦{payment.amount:,.2f} "
            f"for {wallet.student.full_name}",
            request=request, model='Payment', obj_id=payment.id,
        )

    return txn


def reject_payment(payment, rejected_by, reason='', request=None):
    """Reject a pending payment — no wallet change."""
    if payment.status != Payment.PENDING:
        raise ValueError(f"Payment {payment.reference} is already {payment.status}.")

    payment.status         = Payment.REJECTED
    payment.rejection_note = reason
    payment.save(update_fields=['status', 'rejection_note'])

    _audit(
        rejected_by, 'PAYMENT_REJECT',
        f"Rejected payment {payment.reference} ₦{payment.amount:,.2f}. Reason: {reason}",
        request=request, model='Payment', obj_id=payment.id,
    )


# ─────────────────────────────────────────────────────────────────────────────
# 4.  FIFO CHILD SPLIT
# ─────────────────────────────────────────────────────────────────────────────

def _split_payment_fifo(payment, wallet):
    """
    Distribute approved payment amount across ChildBilling rows using FIFO:
    fully pay one child before moving to the next.
    Also updates corresponding FeePayment records.
    """
    # Get all ChildBilling rows for this wallet with remaining balance, oldest first
    children = ChildBilling.objects.filter(
        parent_billing__wallet=wallet,
        parent_billing__session__is_active=True,
    ).select_related('student').order_by('id')

    remaining = payment.amount

    for cb in children:
        if remaining <= 0:
            break
        if cb.balance <= 0:
            continue

        allocated = min(remaining, cb.balance)

        # Record allocation
        ChildPaymentAllocation.objects.create(
            payment=payment,
            child_billing=cb,
            amount_allocated=allocated,
        )

        # Update ChildBilling
        cb.paid += allocated
        cb.save(update_fields=['paid'])

        # Update ParentBilling
        pb = cb.parent_billing
        pb.total_paid += allocated
        pb.save(update_fields=['total_paid'])

        # Update FeePayment records for this student (oldest unpaid first)
        _apply_to_fee_payments(cb.student, allocated, payment)

        remaining -= allocated


def _apply_to_fee_payments(student, amount, payment):
    """Apply amount to student's oldest unpaid FeePayment records."""
    unpaid = FeePayment.objects.filter(
        student=student
    ).exclude(status='PAID').exclude(status='WAIVED').order_by('id')

    remaining = amount
    for fp in unpaid:
        if remaining <= 0:
            break
        owed = fp.balance_due or (fp.fee_structure.amount - fp.amount_paid)
        pay_now = min(remaining, owed)
        fp.amount_paid += pay_now
        fp.balance_due  = max(Decimal('0'), fp.fee_structure.amount - fp.amount_paid)
        fp.status       = 'PAID' if fp.balance_due == 0 else 'PARTIAL'
        fp.paid_at      = timezone.now()
        fp.save(update_fields=['amount_paid', 'balance_due', 'status', 'paid_at'])
        remaining -= pay_now


# ─────────────────────────────────────────────────────────────────────────────
# 5.  PAYSTACK INTEGRATION
# ─────────────────────────────────────────────────────────────────────────────

PAYSTACK_SECRET = getattr(settings, 'PAYSTACK_SECRET_KEY', '')


def initialize_paystack_payment(wallet, amount_naira, email, callback_url):
    """
    Call Paystack API to initialize a transaction.
    Returns (authorization_url, reference) or raises ValueError.
    """
    import requests
    payload = {
        'email':       email,
        'amount':      int(amount_naira * 100),   # kobo
        'reference':   _gen_ref("PSK"),
        'callback_url': callback_url,
        'metadata': {
            'student_id':   wallet.student_id,
            'student_name': wallet.student.full_name,
        },
    }
    try:
        resp = requests.post(
            'https://api.paystack.co/transaction/initialize',
            json=payload,
            headers={
                'Authorization': f'Bearer {PAYSTACK_SECRET}',
                'Content-Type':  'application/json',
            },
            timeout=15,
        )
        data = resp.json()
        if data.get('status'):
            return (
                data['data']['authorization_url'],
                data['data']['reference'],
            )
        raise ValueError(data.get('message', 'Paystack error'))
    except Exception as e:
        raise ValueError(f"Paystack init failed: {e}")


def verify_paystack_signature(request_body: bytes, signature: str) -> bool:
    """Verify that a Paystack webhook came from Paystack."""
    expected = hmac.new(
        PAYSTACK_SECRET.encode(), request_body, hashlib.sha512
    ).hexdigest()
    return hmac.compare_digest(expected, signature)


def handle_paystack_webhook(payload: dict):
    """
    Process a verified Paystack 'charge.success' event.
    Creates a Payment in PENDING state — admin still approves.
    """
    if payload.get('event') != 'charge.success':
        return None

    data      = payload['data']
    ref       = data['reference']
    amount_k  = data['amount']          # kobo
    amount_n  = Decimal(str(amount_k)) / 100
    meta      = data.get('metadata', {})
    student_id = meta.get('student_id')

    # Idempotency — skip if already recorded
    if Payment.objects.filter(paystack_ref=ref).exists():
        return None

    try:
        from users.models import Student as St
        student = St.objects.get(id=student_id)
    except Exception:
        return None

    wallet = Wallet.get_or_create_for_student(student)
    payment = Payment.objects.create(
        wallet=wallet,
        amount=amount_n,
        reference=_gen_ref("PSK"),
        paystack_ref=ref,
        method='PAYSTACK',
        status=Payment.PENDING,
        description=f"Paystack payment {ref}",
    )
    return payment


# ─────────────────────────────────────────────────────────────────────────────
# 6.  PDF RECEIPT  (reportlab platypus)
# ─────────────────────────────────────────────────────────────────────────────

def generate_payment_receipt_pdf(payment):
    """Return BytesIO containing a PDF receipt for an approved payment."""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.lib import colors
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable,
        )
    except ImportError:
        raise ImportError("reportlab is required for PDF generation. pip install reportlab")

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            rightMargin=2*cm, leftMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)

    styles = getSampleStyleSheet()
    green  = colors.HexColor('#064e3b')
    lgray  = colors.HexColor('#f8fafc')
    mgray  = colors.HexColor('#64748b')

    h1 = ParagraphStyle('h1', parent=styles['Normal'],
                        fontSize=20, fontName='Helvetica-Bold',
                        textColor=green, spaceAfter=4)
    normal = ParagraphStyle('n', parent=styles['Normal'],
                            fontSize=10, leading=14)
    small  = ParagraphStyle('s', parent=styles['Normal'],
                            fontSize=8, textColor=mgray)
    center = ParagraphStyle('c', parent=styles['Normal'],
                            fontSize=10, alignment=1)

    wallet  = payment.wallet
    student = wallet.student

    story = [
        Paragraph("WHITE DIAMONDS ACADEMY", h1),
        Paragraph("Zaramangada, Rayfield Road, Jos, Plateau State, Nigeria", small),
        HRFlowable(width="100%", thickness=2, color=green),
        Spacer(1, 0.4*cm),
        Paragraph("<b>OFFICIAL PAYMENT RECEIPT</b>", center),
        Spacer(1, 0.3*cm),
    ]

    data = [
        ["Receipt No.",    payment.reference],
        ["Date",           payment.approved_at.strftime('%d %B %Y %H:%M') if payment.approved_at else '—'],
        ["Student Name",   student.full_name],
        ["Admission No.",  student.admission_number],
        ["Class",          str(student.class_assigned or '—')],
        ["Parent Name",    student.parent_name or '—'],
        ["Amount Paid",    f"₦{payment.amount:,.2f}"],
        ["Payment Method", payment.get_method_display()],
        ["Description",    payment.description or '—'],
        ["Status",         "✓  APPROVED"],
    ]
    t = Table(data, colWidths=[5*cm, 11*cm])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (0,-1), lgray),
        ('FONTNAME',   (0,0), (0,-1), 'Helvetica-Bold'),
        ('FONTNAME',   (1,0), (1,-1), 'Helvetica'),
        ('FONTSIZE',   (0,0), (-1,-1), 10),
        ('GRID',       (0,0), (-1,-1), 0.5, colors.HexColor('#e2e8f0')),
        ('ROWBACKGROUNDS', (0,0), (-1,-1), [colors.white, lgray]),
        ('PADDING',    (0,0), (-1,-1), 8),
        ('TEXTCOLOR',  (1,6), (1,6), green),
        ('FONTNAME',   (1,6), (1,6), 'Helvetica-Bold'),
        ('FONTSIZE',   (1,6), (1,6), 13),
        ('TEXTCOLOR',  (1,9), (1,9), green),
        ('FONTNAME',   (1,9), (1,9), 'Helvetica-Bold'),
    ]))
    story += [t, Spacer(1, 0.6*cm)]

    story += [
        HRFlowable(width="100%", thickness=1, color=colors.HexColor('#e2e8f0'),
                   dash=(3,3)),
        Spacer(1, 0.3*cm),
        Paragraph("This is a computer-generated receipt — no signature required.", small),
        Paragraph(f"© {timezone.now().year} Techmiary Institute of Technology", small),
    ]

    doc.build(story)
    buf.seek(0)
    return buf


# ─────────────────────────────────────────────────────────────────────────────
# 7.  PARENT ACCOUNT STATEMENT  (bank-style with running balance)
# ─────────────────────────────────────────────────────────────────────────────

def build_statement(wallet, from_date=None, to_date=None):
    """
    Returns a list of dicts suitable for template rendering.
    Each row has: date, description, credit, debit, balance.
    Running balance is computed chronologically.
    """
    qs = wallet.transactions.order_by('created_at')
    if from_date:
        qs = qs.filter(created_at__date__gte=from_date)
    if to_date:
        qs = qs.filter(created_at__date__lte=to_date)

    rows = []
    running = Decimal('0')
    for txn in qs:
        if txn.txn_type == Transaction.CREDIT:
            running += txn.amount
            rows.append({
                'date':        txn.created_at,
                'description': txn.description,
                'reference':   txn.reference,
                'credit':      txn.amount,
                'debit':       None,
                'balance':     running,
                'txn':         txn,
            })
        else:
            running -= txn.amount
            rows.append({
                'date':        txn.created_at,
                'description': txn.description,
                'reference':   txn.reference,
                'credit':      None,
                'debit':       txn.amount,
                'balance':     running,
                'txn':         txn,
            })
    return rows


# ─────────────────────────────────────────────────────────────────────────────
# 8.  EXCEL EXPORT  (openpyxl)
# ─────────────────────────────────────────────────────────────────────────────

def export_statement_excel(wallet, from_date=None, to_date=None):
    """Return a BytesIO Excel file containing the wallet statement."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl is required. pip install openpyxl")

    rows  = build_statement(wallet, from_date, to_date)
    wb    = openpyxl.Workbook()
    ws    = wb.active
    ws.title = "Statement"

    # Branding header
    ws.merge_cells('A1:F1')
    ws['A1'] = "WHITE DIAMONDS ACADEMY — WALLET STATEMENT"
    ws['A1'].font      = Font(bold=True, size=14, color="064E3B")
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:F2')
    ws['A2'] = f"Student: {wallet.student.full_name}  |  Admission: {wallet.student.admission_number}  |  Current Balance: ₦{wallet.balance:,.2f}"
    ws['A2'].alignment = Alignment(horizontal='center')
    ws['A2'].font      = Font(size=10, color="475569")

    # Column headers
    headers = ['Date', 'Description', 'Reference', 'Credit (₦)', 'Debit (₦)', 'Balance (₦)']
    hdr_fill = PatternFill(fill_type='solid', fgColor='064E3B')
    hdr_font = Font(bold=True, color='FFFFFF', size=10)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.fill      = hdr_fill
        cell.font      = hdr_font
        cell.alignment = Alignment(horizontal='center')

    # Data rows
    thin = Side(style='thin', color='E2E8F0')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row_idx, r in enumerate(rows, 5):
        ws.cell(row=row_idx, column=1, value=r['date'].strftime('%d %b %Y %H:%M'))
        ws.cell(row=row_idx, column=2, value=r['description'])
        ws.cell(row=row_idx, column=3, value=r['reference'])
        ws.cell(row=row_idx, column=4, value=float(r['credit'])  if r['credit']  else '')
        ws.cell(row=row_idx, column=5, value=float(r['debit'])   if r['debit']   else '')
        ws.cell(row=row_idx, column=6, value=float(r['balance']))
        # Style
        for col in range(1, 7):
            c = ws.cell(row=row_idx, column=col)
            c.border = border
            if row_idx % 2 == 0:
                c.fill = PatternFill(fill_type='solid', fgColor='F0FDF4')
        # Credit green / debit red
        if r['credit']:
            ws.cell(row=row_idx, column=4).font = Font(color='16A34A', bold=True)
        if r['debit']:
            ws.cell(row=row_idx, column=5).font = Font(color='DC2626', bold=True)

    # Column widths
    widths = [20, 40, 18, 14, 14, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_all_billing_excel(session, term):
    """Export all ParentBilling rows for a term to Excel."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl required. pip install openpyxl")

    billings = ParentBilling.objects.filter(
        session=session, term=term
    ).select_related('wallet__student').prefetch_related('children__student')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fee Billing"

    ws.merge_cells('A1:G1')
    ws['A1'] = f"WHITE DIAMONDS ACADEMY — Fee Billing Report  {term} / {session}"
    ws['A1'].font      = Font(bold=True, size=13, color="064E3B")
    ws['A1'].alignment = Alignment(horizontal='center')

    headers = ['Student', 'Admission No.', 'Class', 'Parent Name',
               'Total Fee (₦)', 'Amount Paid (₦)', 'Balance Due (₦)']
    hdr_fill = PatternFill(fill_type='solid', fgColor='1E3A5F')
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.fill = hdr_fill
        c.font = Font(bold=True, color='FFFFFF', size=10)
        c.alignment = Alignment(horizontal='center')

    thin   = Side(style='thin', color='E2E8F0')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    row_n  = 4
    for pb in billings:
        for cb in pb.children.all():
            s = cb.student
            ws.cell(row=row_n, column=1, value=s.full_name)
            ws.cell(row=row_n, column=2, value=s.admission_number)
            ws.cell(row=row_n, column=3, value=str(s.class_assigned or ''))
            ws.cell(row=row_n, column=4, value=s.parent_name or '')
            ws.cell(row=row_n, column=5, value=float(cb.total_fee))
            ws.cell(row=row_n, column=6, value=float(cb.paid))
            bal_cell = ws.cell(row=row_n, column=7, value=float(cb.balance))
            if cb.balance > 0:
                bal_cell.font = Font(color='DC2626', bold=True)
            else:
                bal_cell.font = Font(color='16A34A', bold=True)
            for col in range(1, 8):
                ws.cell(row=row_n, column=col).border = border
                if row_n % 2 == 0:
                    ws.cell(row=row_n, column=col).fill = \
                        PatternFill(fill_type='solid', fgColor='EFF6FF')
            row_n += 1

    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 24
    for col in ['E', 'F', 'G']:
        ws.column_dimensions[col].width = 16

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─────────────────────────────────────────────────────────────────────────────
# 9.  FINANCE ANALYTICS  (chart-ready data for admin dashboard)
# ─────────────────────────────────────────────────────────────────────────────

def get_finance_summary(session=None, term=None):
    """
    Returns a dict of aggregated finance metrics for the dashboard.
    Computed directly from FeeStructure + FeePayment (not ParentBilling)
    so it is always accurate even before parents log in.
    Compatible with Chart.js.
    """
    from django.db.models import Sum, Count, Q, OuterRef, Subquery
    from academics.models import AcademicSession, Term as TermModel
    from users.models import Student as St

    # Defaults
    if not session:
        session = AcademicSession.objects.filter(is_active=True).first()
    if not term and session:
        term = TermModel.objects.filter(session=session, is_active=True).first()

    # ── Total Expected ─────────────────────────────────────────────────────
    # Sum all active FeeStructure amounts applicable to this term/session
    # multiplied by the number of eligible students per fee.
    # Simpler approach: for each active student, sum the fees that apply to them.
    active_students = St.objects.filter(status='Active').select_related('class_assigned')

    total_expected = Decimal('0')
    total_paid     = Decimal('0')
    unpaid_students = []   # list of (student, amount_owed)

    fee_qs = FeeStructure.objects.filter(is_active=True).filter(
        Q(term=term) | Q(term__isnull=True)
    ).filter(
        Q(session=session) | Q(session__isnull=True)
    ) if (session or term) else FeeStructure.objects.filter(is_active=True)

    for student in active_students:
        student_fees = fee_qs.filter(
            Q(school_class=student.class_assigned) | Q(school_class__isnull=True)
        )
        student_expected = sum(f.amount for f in student_fees)
        if student_expected == 0:
            continue

        # Amount paid via FeePayment for fees that apply to this student
        student_paid = FeePayment.objects.filter(
            student=student,
            fee_structure__in=student_fees,
            status__in=['PAID', 'PARTIAL', 'WAIVED'],
        ).aggregate(s=Sum('amount_paid'))['s'] or Decimal('0')

        total_expected += student_expected
        total_paid     += student_paid
        owed = student_expected - student_paid
        if owed > 0:
            unpaid_students.append({
                'student':   student,
                'expected':  student_expected,
                'paid':      student_paid,
                'balance':   owed,
            })

    outstanding     = max(Decimal('0'), total_expected - total_paid)
    collection_rate = (
        round((total_paid / total_expected) * 100, 1)
        if total_expected > 0 else 0
    )

    # Sort unpaid by balance descending, take top 20 for dashboard
    unpaid_students.sort(key=lambda x: x['balance'], reverse=True)

    # ── Pending payments ────────────────────────────────────────────────────
    pending = Payment.objects.filter(status=Payment.PENDING)
    pending_count  = pending.count()
    pending_amount = pending.aggregate(s=Sum('amount'))['s'] or Decimal('0')

    # ── Monthly collection (last 6 months, chart-ready) ────────────────────
    from django.utils import timezone as tz
    from datetime import timedelta
    months_data = []
    now = tz.now()
    for i in range(5, -1, -1):
        month_start = (now.replace(day=1) - timedelta(days=i*28)).replace(day=1)
        month_end   = (month_start.replace(day=28) + timedelta(days=4)).replace(day=1)
        agg_m = Transaction.objects.filter(
            txn_type=Transaction.CREDIT,
            status='COMPLETED',
            created_at__gte=month_start,
            created_at__lt=month_end,
        ).aggregate(s=Sum('amount'))
        months_data.append({
            'month': month_start.strftime('%b %Y'),
            'amount': float(agg_m['s'] or 0),
        })

    # ── Payment method breakdown ────────────────────────────────────────────
    method_breakdown = (
        Payment.objects.filter(status=Payment.APPROVED)
        .values('method')
        .annotate(count=Count('id'), total=Sum('amount'))
    )

    return {
        'total_expected':   total_expected,
        'total_paid':       total_paid,
        'outstanding':      outstanding,
        'collection_rate':  collection_rate,
        'pending_count':    pending_count,
        'pending_amount':   pending_amount,
        'unpaid_students':  unpaid_students[:20],
        'total_unpaid_count': len(unpaid_students),
        'monthly_labels':   [m['month'] for m in months_data],
        'monthly_amounts':  [m['amount'] for m in months_data],
        'method_breakdown': list(method_breakdown),
        'top_outstanding':  unpaid_students[:10],  # kept for template compat
        'session':          session,
        'term':             term,
    }
