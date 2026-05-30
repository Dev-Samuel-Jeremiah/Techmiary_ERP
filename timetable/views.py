from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import IntegrityError
from django.http import HttpResponse
import json

from .models import (
    TimetableConfiguration, TimeSlot, TeacherSubjectClass,
    GeneratedTimetable, TimetableSlot, DayOverride, SpecialSlot,
)
from .forms import TimetableConfigForm, TimeSlotForm, TeacherSubjectClassForm
from .generator import generate_timetable, TimetableGenerationError
from users.models import Class, Staff


def superuser_or_admin_required(view_func):
    """Simple decorator – allow superusers and staff admins."""
    from functools import wraps
    @wraps(view_func)
    def _wrapped(request, *args, **kwargs):
        if not request.user.is_authenticated:
            from django.conf import settings
            return redirect(settings.LOGIN_URL)
        if not (request.user.is_superuser or getattr(request.user, 'is_staff_user', False)):
            messages.error(request, "You don't have permission to access the timetable module.")
            return redirect('dashboard:router')
        return view_func(request, *args, **kwargs)
    return _wrapped


# ---------------------------------------------------------------------------
# Dashboard / Index
# ---------------------------------------------------------------------------

@login_required
def timetable_index(request):
    """
    Public-facing view: shows the published timetable grid for a selected class.
    Also entry point for admin if they have permissions.
    """
    classes = Class.objects.all().order_by('name')
    selected_class_id = request.GET.get('class_id')
    selected_class = None
    timetable_data = None
    days = []
    all_slots = []

    # Get the active published timetable
    published = (
        GeneratedTimetable.objects
        .filter(is_published=True, config__is_active=True)
        .select_related('config')
        .first()
    )

    # Fetch active day overrides for the published config
    day_overrides = {}
    if published:
        for override in (
            published.config.day_overrides
            .filter(is_active=True)
            .prefetch_related('special_slots__supervisor', 'special_slots__classes')
        ):
            day_overrides[override.day] = override

    if published and selected_class_id:
        selected_class = get_object_or_404(Class, pk=selected_class_id)
        config = published.config
        days = [(code, label) for code, label in TimetableConfiguration.DAYS if code in config.active_days]

        # Include ALL slots (breaks + lessons) for the full grid display
        all_slots = list(
            config.time_slots.all().order_by('order', 'start_time')
        )

        # Only fetch lesson data for non-break slots
        slot_data = (
            published.slots
            .filter(school_class=selected_class)
            .select_related('subject', 'teacher', 'time_slot')
        )
        # Build grid: {slot_id: {day_code: TimetableSlot}}
        grid = {}
        for entry in slot_data:
            grid.setdefault(entry.time_slot_id, {})[entry.day] = entry

        # Build timetable_data with per-day override awareness
        # Each row: (slot, day_cell_list)
        # day_cell_list: list of dicts with keys: code, label, entry, cell_type
        #   cell_type: 'lesson' | 'break' | 'early_close' | 'special'
        timetable_rows = []
        for slot in all_slots:
            if slot.is_break:
                # Break row — span all days
                timetable_rows.append({
                    'slot': slot,
                    'is_break': True,
                    'day_cells': None,
                })
            else:
                day_cells = []
                for code, label in days:
                    override = day_overrides.get(code)
                    cell = {'code': code, 'label': label}
                    if override and override.is_active and slot.start_time >= override.lessons_end_time:
                        # This slot falls after the override cutoff — check for special slot
                        special = None
                        for ss in override.special_slots.all():
                            if ss.start_time <= slot.start_time < ss.end_time:
                                special = ss
                                break
                        if special:
                            cell['cell_type'] = 'special'
                            cell['special_slot'] = special
                        else:
                            cell['cell_type'] = 'early_close'
                            cell['override_label'] = override.label
                    else:
                        entry = grid.get(slot.pk, {}).get(code)
                        cell['cell_type'] = 'lesson'
                        cell['entry'] = entry
                    day_cells.append(cell)
                timetable_rows.append({
                    'slot': slot,
                    'is_break': False,
                    'day_cells': day_cells,
                })

        timetable_data = timetable_rows

    return render(request, 'timetable/timetable_view.html', {
        'classes': classes,
        'selected_class': selected_class,
        'published': published,
        'timetable_data': timetable_data,
        'days': days,
        'all_slots': all_slots,
        'day_overrides': day_overrides,
    })


@login_required
def teacher_timetable(request):
    """Teacher's personal timetable view."""
    user = request.user
    teacher_staff = getattr(user, 'staff', None)

    published = (
        GeneratedTimetable.objects
        .filter(is_published=True, config__is_active=True)
        .select_related('config')
        .first()
    )

    timetable_data = None
    days = []
    slots = []

    if published and teacher_staff:
        config = published.config
        days = [(code, label) for code, label in TimetableConfiguration.DAYS if code in config.active_days]
        slots = list(config.get_active_slots())
        slot_data = (
            published.slots
            .filter(teacher=teacher_staff)
            .select_related('subject', 'school_class', 'time_slot')
        )
        grid = {}
        for entry in slot_data:
            grid.setdefault(entry.time_slot_id, {})[entry.day] = entry
        timetable_data = [(slot, grid.get(slot.pk, {})) for slot in slots]

    return render(request, 'timetable/teacher_timetable.html', {
        'published': published,
        'timetable_data': timetable_data,
        'days': days,
        'slots': slots,
        'teacher_staff': teacher_staff,
    })


# ---------------------------------------------------------------------------
# Admin: Configuration
# ---------------------------------------------------------------------------

@superuser_or_admin_required
def admin_index(request):
    configs = TimetableConfiguration.objects.prefetch_related('time_slots').all()
    return render(request, 'timetable/admin/index.html', {'configs': configs})


@superuser_or_admin_required
def config_create(request):
    if request.method == 'POST':
        form = TimetableConfigForm(request.POST)
        if form.is_valid():
            obj = form.save()
            messages.success(request, f'Configuration "{obj.name}" created successfully.')
            return redirect('timetable:config_detail', pk=obj.pk)
    else:
        form = TimetableConfigForm()
    time_slots = TimeSlot.objects.all().order_by('order')
    return render(request, 'timetable/admin/config_form.html', {
        'form': form, 'time_slots': time_slots, 'action': 'Create'
    })


@superuser_or_admin_required
def config_edit(request, pk):
    config = get_object_or_404(TimetableConfiguration, pk=pk)
    if request.method == 'POST':
        form = TimetableConfigForm(request.POST, instance=config)
        if form.is_valid():
            obj = form.save()
            messages.success(request, 'Configuration updated.')
            return redirect('timetable:config_detail', pk=obj.pk)
    else:
        form = TimetableConfigForm(instance=config)
        # Pre-fill active_days_select
        form.fields['active_days_select'].initial = config.active_days
    return render(request, 'timetable/admin/config_form.html', {
        'form': form, 'config': config, 'action': 'Edit'
    })


@superuser_or_admin_required
def config_detail(request, pk):
    config = get_object_or_404(TimetableConfiguration, pk=pk)
    assignments = config.teacher_subject_classes.select_related('staff', 'subject', 'school_class').all()
    generated = config.generated_timetables.all()[:5]
    return render(request, 'timetable/admin/config_detail.html', {
        'config': config,
        'assignments': assignments,
        'generated': generated,
    })


@superuser_or_admin_required
def config_delete(request, pk):
    config = get_object_or_404(TimetableConfiguration, pk=pk)
    if request.method == 'POST':
        name = config.name
        config.delete()
        messages.success(request, f'Configuration "{name}" deleted.')
        return redirect('timetable:admin_index')
    return render(request, 'timetable/admin/confirm_delete.html', {
        'object': config, 'object_type': 'Configuration'
    })


# ---------------------------------------------------------------------------
# Admin: Time Slots
# ---------------------------------------------------------------------------

@superuser_or_admin_required
def slot_list(request):
    slots = TimeSlot.objects.all()
    return render(request, 'timetable/admin/slot_list.html', {'slots': slots})


@superuser_or_admin_required
def slot_create(request):
    if request.method == 'POST':
        form = TimeSlotForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Time slot created.')
            return redirect('timetable:slot_list')
    else:
        form = TimeSlotForm()
    return render(request, 'timetable/admin/slot_form.html', {'form': form, 'action': 'Create'})


@superuser_or_admin_required
def slot_edit(request, pk):
    slot = get_object_or_404(TimeSlot, pk=pk)
    if request.method == 'POST':
        form = TimeSlotForm(request.POST, instance=slot)
        if form.is_valid():
            form.save()
            messages.success(request, 'Time slot updated.')
            return redirect('timetable:slot_list')
    else:
        form = TimeSlotForm(instance=slot)
    return render(request, 'timetable/admin/slot_form.html', {'form': form, 'slot': slot, 'action': 'Edit'})


@superuser_or_admin_required
def slot_delete(request, pk):
    slot = get_object_or_404(TimeSlot, pk=pk)
    if request.method == 'POST':
        slot.delete()
        messages.success(request, 'Time slot deleted.')
        return redirect('timetable:slot_list')
    return render(request, 'timetable/admin/confirm_delete.html', {
        'object': slot, 'object_type': 'Time Slot'
    })


# ---------------------------------------------------------------------------
# Admin: Teacher-Subject-Class Assignments
# ---------------------------------------------------------------------------

@superuser_or_admin_required
def assignment_create(request, config_pk):
    config = get_object_or_404(TimetableConfiguration, pk=config_pk)
    if request.method == 'POST':
        form = TeacherSubjectClassForm(request.POST, config=config)
        if form.is_valid():
            try:
                form.save()
                messages.success(request, 'Assignment added successfully.')
            except IntegrityError:
                messages.error(request, 'This assignment already exists.')
            return redirect('timetable:config_detail', pk=config_pk)
    else:
        form = TeacherSubjectClassForm(config=config)
    return render(request, 'timetable/admin/assignment_form.html', {
        'form': form, 'config': config, 'action': 'Add'
    })


@superuser_or_admin_required
def assignment_edit(request, pk):
    assignment = get_object_or_404(TeacherSubjectClass, pk=pk)
    config = assignment.config
    if request.method == 'POST':
        form = TeacherSubjectClassForm(request.POST, instance=assignment, config=config)
        if form.is_valid():
            form.save()
            messages.success(request, 'Assignment updated.')
            return redirect('timetable:config_detail', pk=config.pk)
    else:
        form = TeacherSubjectClassForm(instance=assignment, config=config)
    return render(request, 'timetable/admin/assignment_form.html', {
        'form': form, 'config': config, 'assignment': assignment, 'action': 'Edit'
    })


@superuser_or_admin_required
def assignment_delete(request, pk):
    assignment = get_object_or_404(TeacherSubjectClass, pk=pk)
    config_pk = assignment.config_id
    if request.method == 'POST':
        assignment.delete()
        messages.success(request, 'Assignment removed.')
        return redirect('timetable:config_detail', pk=config_pk)
    return render(request, 'timetable/admin/confirm_delete.html', {
        'object': assignment, 'object_type': 'Assignment'
    })


# ---------------------------------------------------------------------------
# Admin: Generate & Publish
# ---------------------------------------------------------------------------

@superuser_or_admin_required
def generate(request, config_pk):
    config = get_object_or_404(TimetableConfiguration, pk=config_pk)
    if request.method == 'POST':
        try:
            timetable = generate_timetable(config.pk, max_attempts=30, timeout_seconds=20)
            messages.success(
                request,
                f'Timetable generated successfully with {timetable.slots.count()} slots.'
                ' You can now preview and publish it.'
            )
            return redirect('timetable:preview', pk=timetable.pk)
        except TimetableGenerationError as e:
            messages.error(request, str(e))
            return redirect('timetable:generate', config_pk=config_pk)

    # Build capacity check data for the confirmation page
    slots = list(config.get_active_slots())
    days = config.active_days
    capacity = len(days) * len(slots)

    from collections import defaultdict
    assignments = list(
        config.teacher_subject_classes.select_related('staff', 'subject', 'school_class')
    )
    class_demand = defaultdict(int)
    class_names = {}
    for a in assignments:
        class_demand[a.school_class.pk] += a.periods_per_week
        class_names[a.school_class.pk] = a.school_class.name

    capacity_rows = []
    has_overload = False
    for cid, demand in class_demand.items():
        ok = demand <= capacity
        if not ok:
            has_overload = True
        capacity_rows.append({
            'name': class_names[cid],
            'demand': demand,
            'capacity': capacity,
            'ok': ok,
            'over': demand - capacity if not ok else 0,
        })
    capacity_rows.sort(key=lambda r: r['name'])

    return render(request, 'timetable/admin/generate_confirm.html', {
        'config': config,
        'capacity': capacity,
        'capacity_rows': capacity_rows,
        'has_overload': has_overload,
        'assignments_by_class': bool(class_demand),
    })


@superuser_or_admin_required
def preview(request, pk):
    timetable = get_object_or_404(GeneratedTimetable, pk=pk)
    config = timetable.config
    classes = Class.objects.filter(
        pk__in=timetable.slots.values_list('school_class', flat=True).distinct()
    ).order_by('name')
    days = [(code, label) for code, label in TimetableConfiguration.DAYS if code in config.active_days]
    all_slots = list(config.time_slots.all().order_by('order', 'start_time'))

    day_overrides = {}
    for override in (
        config.day_overrides
        .filter(is_active=True)
        .prefetch_related('special_slots__supervisor', 'special_slots__classes')
    ):
        day_overrides[override.day] = override

    # Build per-class grid
    class_grids = {}
    for cls in classes:
        slot_data = (
            timetable.slots
            .filter(school_class=cls)
            .select_related('subject', 'teacher', 'time_slot')
        )
        grid = {}
        for entry in slot_data:
            grid.setdefault(entry.time_slot_id, {})[entry.day] = entry

        rows = []
        for slot in all_slots:
            if slot.is_break:
                rows.append({'slot': slot, 'is_break': True, 'day_cells': None})
            else:
                day_cells = []
                for code, label in days:
                    override = day_overrides.get(code)
                    cell = {'code': code, 'label': label}
                    if override and override.is_active and slot.start_time >= override.lessons_end_time:
                        special = None
                        for ss in override.special_slots.all():
                            if ss.start_time <= slot.start_time < ss.end_time:
                                special = ss
                                break
                        if special:
                            cell['cell_type'] = 'special'
                            cell['special_slot'] = special
                        else:
                            cell['cell_type'] = 'early_close'
                            cell['override_label'] = override.label
                    else:
                        entry = grid.get(slot.pk, {}).get(code)
                        cell['cell_type'] = 'lesson'
                        cell['entry'] = entry
                    day_cells.append(cell)
                rows.append({'slot': slot, 'is_break': False, 'day_cells': day_cells})
        class_grids[cls] = rows

    return render(request, 'timetable/admin/preview.html', {
        'timetable': timetable,
        'config': config,
        'classes': classes,
        'class_grids': class_grids,
        'days': days,
        'all_slots': all_slots,
        'day_overrides': day_overrides,
    })


@superuser_or_admin_required
def publish(request, pk):
    timetable = get_object_or_404(GeneratedTimetable, pk=pk)
    if request.method == 'POST':
        timetable.publish()
        messages.success(request, 'Timetable published! Students and teachers can now view it.')
        return redirect('timetable:preview', pk=pk)
    return render(request, 'timetable/admin/confirm_publish.html', {'timetable': timetable})


@superuser_or_admin_required
def delete_generated(request, pk):
    timetable = get_object_or_404(GeneratedTimetable, pk=pk)
    config_pk = timetable.config_id
    if request.method == 'POST':
        timetable.delete()
        messages.success(request, 'Generated timetable deleted.')
        return redirect('timetable:config_detail', pk=config_pk)
    return render(request, 'timetable/admin/confirm_delete.html', {
        'object': timetable, 'object_type': 'Generated Timetable'
    })


# ---------------------------------------------------------------------------
# Admin: Day Overrides (e.g. Friday Computer Club)
# ---------------------------------------------------------------------------

@superuser_or_admin_required
def day_override_list(request, config_pk):
    config = get_object_or_404(TimetableConfiguration, pk=config_pk)
    overrides = config.day_overrides.prefetch_related('special_slots').all()
    return render(request, 'timetable/admin/day_override_list.html', {
        'config': config,
        'overrides': overrides,
    })


@superuser_or_admin_required
def day_override_create(request, config_pk):
    config = get_object_or_404(TimetableConfiguration, pk=config_pk)
    if request.method == 'POST':
        day = request.POST.get('day')
        label = request.POST.get('label', '').strip()
        description = request.POST.get('description', '').strip()
        lessons_end_time = request.POST.get('lessons_end_time')
        is_active = request.POST.get('is_active') == 'on'

        if not day or not label or not lessons_end_time:
            messages.error(request, 'Day, label, and lessons end time are required.')
        else:
            try:
                override = DayOverride.objects.create(
                    config=config,
                    day=day,
                    label=label,
                    description=description,
                    lessons_end_time=lessons_end_time,
                    is_active=is_active,
                )
                messages.success(request, f'Day override "{override.label}" created.')
                return redirect('timetable:day_override_detail', config_pk=config_pk, pk=override.pk)
            except Exception as e:
                messages.error(request, f'Error: {e}')

    available_days = [
        (code, label_) for code, label_ in TimetableConfiguration.DAYS
        if code in config.active_days
        and not config.day_overrides.filter(day=code).exists()
    ]
    return render(request, 'timetable/admin/day_override_form.html', {
        'config': config,
        'available_days': available_days,
        'action': 'Create',
    })


@superuser_or_admin_required
def day_override_detail(request, config_pk, pk):
    config = get_object_or_404(TimetableConfiguration, pk=config_pk)
    override = get_object_or_404(DayOverride, pk=pk, config=config)
    special_slots = override.get_special_slots()
    all_staff = Staff.objects.all().order_by('user__last_name')
    all_classes = Class.objects.all().order_by('name')
    return render(request, 'timetable/admin/day_override_detail.html', {
        'config': config,
        'override': override,
        'special_slots': special_slots,
        'all_staff': all_staff,
        'all_classes': all_classes,
    })


@superuser_or_admin_required
def day_override_edit(request, config_pk, pk):
    config = get_object_or_404(TimetableConfiguration, pk=config_pk)
    override = get_object_or_404(DayOverride, pk=pk, config=config)
    if request.method == 'POST':
        override.label = request.POST.get('label', override.label).strip()
        override.description = request.POST.get('description', '').strip()
        override.lessons_end_time = request.POST.get('lessons_end_time', override.lessons_end_time)
        override.is_active = request.POST.get('is_active') == 'on'
        try:
            override.save()
            messages.success(request, 'Day override updated.')
            return redirect('timetable:day_override_detail', config_pk=config_pk, pk=pk)
        except Exception as e:
            messages.error(request, f'Error: {e}')
    return render(request, 'timetable/admin/day_override_form.html', {
        'config': config,
        'override': override,
        'action': 'Edit',
    })


@superuser_or_admin_required
def day_override_delete(request, config_pk, pk):
    config = get_object_or_404(TimetableConfiguration, pk=config_pk)
    override = get_object_or_404(DayOverride, pk=pk, config=config)
    if request.method == 'POST':
        override.delete()
        messages.success(request, 'Day override deleted.')
        return redirect('timetable:day_override_list', config_pk=config_pk)
    return render(request, 'timetable/admin/confirm_delete.html', {
        'object': override,
        'object_type': 'Day Override',
    })


@superuser_or_admin_required
def special_slot_create(request, override_pk):
    override = get_object_or_404(DayOverride, pk=override_pk)
    config_pk = override.config_id
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        slot_type = request.POST.get('slot_type', 'CLUB')
        start_time = request.POST.get('start_time')
        end_time = request.POST.get('end_time')
        venue = request.POST.get('venue', '').strip()
        supervisor_id = request.POST.get('supervisor') or None
        notes = request.POST.get('notes', '').strip()
        applies_to_all = request.POST.get('applies_to_all_classes') == 'on'
        class_ids = request.POST.getlist('classes')

        if not name or not start_time or not end_time:
            messages.error(request, 'Name, start time, and end time are required.')
        else:
            try:
                ss = SpecialSlot.objects.create(
                    override=override,
                    name=name,
                    slot_type=slot_type,
                    start_time=start_time,
                    end_time=end_time,
                    venue=venue,
                    supervisor_id=supervisor_id,
                    notes=notes,
                    applies_to_all_classes=applies_to_all,
                )
                if not applies_to_all and class_ids:
                    ss.classes.set(class_ids)
                messages.success(request, f'Special slot "{ss.name}" added.')
                return redirect('timetable:day_override_detail', config_pk=config_pk, pk=override_pk)
            except Exception as e:
                messages.error(request, f'Error: {e}')

    all_staff = Staff.objects.all().order_by('user__last_name')
    all_classes = Class.objects.all().order_by('name')
    return render(request, 'timetable/admin/special_slot_form.html', {
        'override': override,
        'config_pk': config_pk,
        'all_staff': all_staff,
        'all_classes': all_classes,
        'slot_types': SpecialSlot.SLOT_TYPE_CHOICES,
        'action': 'Add',
    })


@superuser_or_admin_required
def special_slot_edit(request, pk):
    ss = get_object_or_404(SpecialSlot, pk=pk)
    override = ss.override
    config_pk = override.config_id
    if request.method == 'POST':
        ss.name = request.POST.get('name', ss.name).strip()
        ss.slot_type = request.POST.get('slot_type', ss.slot_type)
        ss.start_time = request.POST.get('start_time', ss.start_time)
        ss.end_time = request.POST.get('end_time', ss.end_time)
        ss.venue = request.POST.get('venue', '').strip()
        ss.supervisor_id = request.POST.get('supervisor') or None
        ss.notes = request.POST.get('notes', '').strip()
        ss.applies_to_all_classes = request.POST.get('applies_to_all_classes') == 'on'
        class_ids = request.POST.getlist('classes')
        try:
            ss.full_clean()
            ss.save()
            if not ss.applies_to_all_classes:
                ss.classes.set(class_ids)
            else:
                ss.classes.clear()
            messages.success(request, f'Special slot "{ss.name}" updated.')
            return redirect('timetable:day_override_detail', config_pk=config_pk, pk=override.pk)
        except Exception as e:
            messages.error(request, f'Error: {e}')

    all_staff = Staff.objects.all().order_by('user__last_name')
    all_classes = Class.objects.all().order_by('name')
    return render(request, 'timetable/admin/special_slot_form.html', {
        'override': override,
        'special_slot': ss,
        'config_pk': config_pk,
        'all_staff': all_staff,
        'all_classes': all_classes,
        'slot_types': SpecialSlot.SLOT_TYPE_CHOICES,
        'action': 'Edit',
    })


@superuser_or_admin_required
def special_slot_delete(request, pk):
    ss = get_object_or_404(SpecialSlot, pk=pk)
    override = ss.override
    config_pk = override.config_id
    if request.method == 'POST':
        name = ss.name
        ss.delete()
        messages.success(request, f'Special slot "{name}" removed.')
        return redirect('timetable:day_override_detail', config_pk=config_pk, pk=override.pk)
    return render(request, 'timetable/admin/confirm_delete.html', {
        'object': ss,
        'object_type': 'Special Slot',
    })


# ---------------------------------------------------------------------------
# Excel Export — all classes, one sheet per class
# ---------------------------------------------------------------------------

def _fmt12(t):
    """Convert a time object to 12-hour AM/PM string."""
    if t is None:
        return ''
    h, m = t.hour, t.minute
    period = 'AM' if h < 12 else 'PM'
    return f"{h % 12 or 12}:{m:02d} {period}"


@login_required
def export_excel(request, pk):
    """
    Export the generated timetable to a styled .xlsx file.
    One worksheet per class.  Accessible to admins and superusers.
    """
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, GradientFill
    )
    from openpyxl.utils import get_column_letter

    if not (request.user.is_superuser or getattr(request.user, 'is_staff_user', False)):
        messages.error(request, "You don't have permission to export timetables.")
        return redirect('dashboard:router')

    timetable = get_object_or_404(GeneratedTimetable, pk=pk)
    config    = timetable.config
    days      = [(code, label) for code, label in TimetableConfiguration.DAYS
                 if code in config.active_days]

    # All slots including breaks
    all_slots = list(config.time_slots.all().order_by('order', 'start_time'))

    # Day overrides
    day_overrides = {}
    for override in config.day_overrides.filter(is_active=True).prefetch_related(
        'special_slots__supervisor', 'special_slots__classes'
    ):
        day_overrides[override.day] = override

    # Classes that have slots in this timetable
    classes = Class.objects.filter(
        pk__in=timetable.slots.values_list('school_class', flat=True).distinct()
    ).order_by('name')

    # ------------------------------------------------------------------ #
    # Styles                                                               #
    # ------------------------------------------------------------------ #
    DARK      = '0F1724'
    DARK2     = '1E293B'
    AMBER     = '78350F'
    BREAK_BG  = 'C7D2FE'   # indigo-200
    BREAK_FG  = '3730A3'
    LESSON_BG = 'EFF6FF'   # blue-50
    LESSON_BD = 'BFDBFE'   # blue-200
    EARLY_BG  = 'FFFBEB'
    EARLY_FG  = 'D97706'
    SPEC_COLORS = {
        'CLUB':     ('FFFBEB', 'D97706'),
        'SPORT':    ('F0FDF4', '059669'),
        'EXAM':     ('FFF1F2', 'DC2626'),
        'ASSEMBLY': ('F0F4FF', '4F46E5'),
        'OTHER':    ('FAF5FF', '7C3AED'),
    }

    thin = Side(style='thin', color='D1D5DB')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left   = Alignment(horizontal='left',   vertical='center', wrap_text=True)

    def hdr_cell(ws, row, col, value, bg=DARK, fg='FFFFFF', bold=True, size=11):
        c = ws.cell(row=row, column=col, value=value)
        c.font      = Font(name='Arial', bold=bold, color=fg, size=size)
        c.fill      = PatternFill('solid', start_color=bg)
        c.alignment = center
        c.border    = border
        return c

    def data_cell(ws, row, col, value, bg='FFFFFF', fg='1E293B',
                  bold=False, size=10, align=center):
        c = ws.cell(row=row, column=col, value=value)
        c.font      = Font(name='Arial', bold=bold, color=fg, size=size)
        c.fill      = PatternFill('solid', start_color=bg)
        c.alignment = align
        c.border    = border
        return c

    # ------------------------------------------------------------------ #
    # Build workbook                                                       #
    # ------------------------------------------------------------------ #
    wb = Workbook()
    wb.remove(wb.active)   # remove default blank sheet

    for cls in classes:
        ws = wb.create_sheet(title=cls.name[:31])   # sheet names max 31 chars

        # Fetch this class's slot data
        slot_data = (
            timetable.slots
            .filter(school_class=cls)
            .select_related('subject', 'teacher', 'time_slot')
        )
        grid = {}
        for entry in slot_data:
            grid.setdefault(entry.time_slot_id, {})[entry.day] = entry

        n_days = len(days)
        # Columns: col1=Period/Time, then one col per day
        # Row 1: big title, Row 2: day headers, Row 3+: slot rows

        # --- Row 1: Title ---
        ws.merge_cells(start_row=1, start_column=1,
                       end_row=1,   end_column=1 + n_days)
        hdr_cell(ws, 1, 1,
                 f"TIMETABLE — {cls.name.upper()}  |  {config.name}",
                 bg=DARK, size=13)

        # --- Row 2: column headers ---
        hdr_cell(ws, 2, 1, 'Period / Time', bg=DARK2, size=10)
        for ci, (code, label) in enumerate(days, start=2):
            override = day_overrides.get(code)
            bg = AMBER if override else DARK2
            txt = label.upper()
            if override:
                txt += f"\n(closes {_fmt12(override.lessons_end_time)})"
            hdr_cell(ws, 2, ci, txt, bg=bg, size=10)
        ws.row_dimensions[2].height = 30

        # --- Data rows ---
        cur_row = 3
        for slot in all_slots:
            if slot.is_break:
                # Merge entire row
                ws.merge_cells(start_row=cur_row, start_column=1,
                                end_row=cur_row,   end_column=1 + n_days)
                c = ws.cell(
                    row=cur_row, column=1,
                    value=f"☕  {slot.name}  ·  {_fmt12(slot.start_time)} – {_fmt12(slot.end_time)}"
                )
                c.font      = Font(name='Arial', bold=True, color=BREAK_FG, size=10)
                c.fill      = PatternFill('solid', start_color=BREAK_BG)
                c.alignment = center
                c.border    = border
                ws.row_dimensions[cur_row].height = 18
            else:
                period_txt = f"{slot.name}\n{_fmt12(slot.start_time)}–{_fmt12(slot.end_time)}"
                data_cell(ws, cur_row, 1, period_txt,
                          bg='F1F5F9', fg='374151', bold=True, size=9, align=center)

                for ci, (code, label) in enumerate(days, start=2):
                    override = day_overrides.get(code)
                    if override and override.is_active and slot.start_time >= override.lessons_end_time:
                        # Check for a special slot covering this time
                        special = None
                        for ss in override.special_slots.all():
                            if ss.start_time <= slot.start_time < ss.end_time:
                                special = ss
                                break
                        if special:
                            bg, fg = SPEC_COLORS.get(special.slot_type, ('FAF5FF', '7C3AED'))
                            txt = f"{special.name}\n{_fmt12(special.start_time)}–{_fmt12(special.end_time)}"
                            data_cell(ws, cur_row, ci, txt, bg=bg, fg=fg, bold=True, size=9)
                        else:
                            ol = override.label
                            data_cell(ws, cur_row, ci, ol,
                                      bg=EARLY_BG, fg=EARLY_FG, bold=True, size=9)
                    else:
                        entry = grid.get(slot.pk, {}).get(code)
                        if entry:
                            txt = f"{entry.subject.name}\n{entry.teacher.full_name}"
                            data_cell(ws, cur_row, ci, txt,
                                      bg=LESSON_BG, fg='1E293B', bold=False, size=9)
                        else:
                            data_cell(ws, cur_row, ci, '—',
                                      bg='FAFAFA', fg='D1D5DB', size=10)
                ws.row_dimensions[cur_row].height = 32

            cur_row += 1

        # --- Column widths ---
        ws.column_dimensions[get_column_letter(1)].width = 16
        for ci in range(2, 2 + n_days):
            ws.column_dimensions[get_column_letter(ci)].width = 20

        # Freeze panes below header rows
        ws.freeze_panes = 'A3'

    # ------------------------------------------------------------------ #
    # Stream the response                                                  #
    # ------------------------------------------------------------------ #
    from io import BytesIO
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_name = config.name.replace(' ', '_').replace('/', '-')[:40]
    filename  = f"Timetable_{safe_name}.xlsx"

    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
